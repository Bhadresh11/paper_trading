# ema_trailing_live_final_v2.py
import websocket
import json
import pandas as pd
import requests
from datetime import datetime
import pytz
import os
from openpyxl import load_workbook
import time
import traceback

# ===========================================================
#                 USER CONFIGURATION PANEL
# ===========================================================
SYMBOL = "BTCUSDT"
INTERVAL = "3m"
HISTORY_LIMIT = 500

capital_start = 100000.0

# --- EMA SETTINGS (change here) ---
EMA_FAST = 5
EMA_SLOW = 20

# --- RISK SETTINGS (0-100, user-friendly) ---
stop_loss_pct = 2        # percent (if you want percent-based SL; we use point-based TSL primarily)
target_pct = 1           # percent TP
risk_per_trade_pct = 1   # percent of capital used for pnl scaling

# --- TRAILING STOP SETTINGS (points) ---
TSL_DISTANCE = 500.0     # points distance (point-based trailing stop)

# -----------------------------------------------------------
# Internal derived factors (converted from user percentages)
sl_factor = stop_loss_pct / 100.0
tp_factor = target_pct / 100.0
risk_factor = risk_per_trade_pct / 100.0

ts = datetime.now().strftime("%m%d_%H%M%S")
EXCEL_FILE = f"excel_data/{EMA_SLOW}_{EMA_FAST}_{ts}.xlsx"

TZ_BKK = pytz.timezone("Asia/Bangkok")
# ===========================================================

# ---------- In-memory data ----------
candles = pd.DataFrame(columns=[
    'open','high','low','close','volume','candle_time_bkk','EMA_fast','EMA_slow'
])
history_crosses = []   # detected crosses on history (list of dicts)
hist_trades_df = pd.DataFrame()  # historical backtest results
swings_for_file = pd.DataFrame()
trades = []            # executed trades (historical closed + live)
pending_order = None   # pending order to execute at next candle open
capital = capital_start

# ---------- Helper utilities ----------
def round2(v):
    return round(float(v), 2)

def to_bkk(ts_ms):
    return datetime.fromtimestamp(ts_ms/1000, tz=pytz.utc).astimezone(TZ_BKK)

def to_bkk_str(ts_ms):
    return to_bkk(ts_ms).strftime("%Y-%m-%d %H:%M:%S")

def binance_klines(symbol, interval, limit=500):
    url = "https://api.binance.com/api/v3/klines"
    params = {"symbol": symbol, "interval": interval, "limit": limit}
    r = requests.get(url, params=params, timeout=10)
    r.raise_for_status()
    return r.json()

def calculate_emas(df):
    if len(df) >= 1:
        df['EMA_fast'] = df['close'].ewm(span=EMA_FAST, adjust=False).mean()
        df['EMA_slow'] = df['close'].ewm(span=EMA_SLOW, adjust=False).mean()
    return df

def detect_crosses(df):
    crosses = []
    if 'EMA_fast' not in df or 'EMA_slow' not in df:
        return crosses
    for i in range(1, len(df)):
        a = df.iloc[i-1]
        b = df.iloc[i]
        if pd.isna(a['EMA_fast']) or pd.isna(a['EMA_slow']) or pd.isna(b['EMA_fast']) or pd.isna(b['EMA_slow']):
            continue
        if (a['EMA_fast'] <= a['EMA_slow']) and (b['EMA_fast'] > b['EMA_slow']):
            crosses.append({"index": i, "type": "GOLDEN", "time_bkk": b['candle_time_bkk'], "price": b['close']})
        if (a['EMA_fast'] >= a['EMA_slow']) and (b['EMA_fast'] < b['EMA_slow']):
            crosses.append({"index": i, "type": "DEAD", "time_bkk": b['candle_time_bkk'], "price": b['close']})
    return crosses

def ensure_excel_run_sheets(base_file, run_tag):
    """Create timestamped sheets for this run and return names."""
    sheet_candles = f"Candles_{run_tag}"
    sheet_swings = f"Swings_{run_tag}"
    sheet_trades = f"Trades_{run_tag}"
    if not os.path.exists(base_file):
        writer = pd.ExcelWriter(base_file, engine='openpyxl', mode='w')
        pd.DataFrame().to_excel(writer, sheet_name=sheet_candles, index=False)
        pd.DataFrame().to_excel(writer, sheet_name=sheet_swings, index=False)
        pd.DataFrame().to_excel(writer, sheet_name=sheet_trades, index=False)
        writer.close()
    else:
        wb = load_workbook(base_file)
        def unique(name):
            if name not in wb.sheetnames:
                return name
            n = 1
            while f"{name}_{n}" in wb.sheetnames:
                n += 1
            return f"{name}_{n}"
        sheet_candles = unique(sheet_candles)
        sheet_swings = unique(sheet_swings)
        sheet_trades = unique(sheet_trades)
        writer = pd.ExcelWriter(base_file, engine='openpyxl', mode='a')
        pd.DataFrame().to_excel(writer, sheet_name=sheet_candles, index=False)
        pd.DataFrame().to_excel(writer, sheet_name=sheet_swings, index=False)
        pd.DataFrame().to_excel(writer, sheet_name=sheet_trades, index=False)
        writer.close()
    return sheet_candles, sheet_swings, sheet_trades

def write_df_safe(base_file, sheet_name, df):
    """Write DataFrame to Excel; make datetimes naive and round floats."""
    df_copy = df.copy()
    # Convert timezone-aware datetimes to naive for Excel
    for col in df_copy.columns:
        if 'time' in col and col in df_copy and pd.api.types.is_datetime64_any_dtype(df_copy[col]):
            df_copy[col] = df_copy[col].dt.tz_convert(None).dt.tz_localize(None) if hasattr(df_copy[col].dt, 'tz_convert') else df_copy[col].dt.tz_localize(None)
    # Convert object datetimes (strings) are okay; ensure numeric rounding
    for c in df_copy.columns:
        if pd.api.types.is_float_dtype(df_copy[c]) or pd.api.types.is_integer_dtype(df_copy[c]):
            try:
                df_copy[c] = df_copy[c].round(2)
            except Exception:
                pass
    # save
    with pd.ExcelWriter(base_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_copy.to_excel(writer, sheet_name=sheet_name, index=False)

# ---------- Historical preload & backtest with Option A TSL (point-based) ----------
print("Fetching historical candles...")
klines = binance_klines(SYMBOL, INTERVAL, HISTORY_LIMIT)
rows = []
for k in klines:
    rows.append({
        "open": round2(k[1]),
        "high": round2(k[2]),
        "low": round2(k[3]),
        "close": round2(k[4]),
        "volume": round2(k[5]),
        "candle_time_bkk": to_bkk(k[0])
    })
candles = pd.DataFrame(rows)
candles = calculate_emas(candles)
print(f"Loaded {len(candles)} historical candles.")

history_crosses = detect_crosses(candles)
print(f"Detected {len(history_crosses)} historical cross events.")

def backtest_historical_with_tsl(candles_df, crosses, tsl_distance, capital_start, risk_factor):
    capital = capital_start
    hist_trades = []
    for cross in crosses:
        entry_idx = cross['index']
        entry_price = candles_df.iloc[entry_idx]['close']
        entry_time = candles_df.iloc[entry_idx]['candle_time_bkk']
        side = "LONG" if cross['type'] == "GOLDEN" else "SHORT"

        # initialize
        if side == "LONG":
            highest = entry_price
            tsl = highest - tsl_distance
            sl = entry_price - tsl_distance
            tp = entry_price * (1 + tp_factor)
        else:
            lowest = entry_price
            tsl = lowest + tsl_distance
            sl = entry_price + tsl_distance
            tp = entry_price * (1 - tp_factor)

        exit_price = None
        exit_time = None
        exit_reason = None

        # step forward to find exit
        for j in range(entry_idx + 1, len(candles_df)):
            r = candles_df.iloc[j]
            high = r['high']; low = r['low']; close = r['close']

            if side == "LONG":
                # update highest and tsl only when price makes new high
                if high > highest:
                    highest = high
                    tsl = highest - tsl_distance
                # TP
                if high >= tp:
                    exit_price = tp
                    exit_time = r['candle_time_bkk']
                    exit_reason = 'TP'
                    break
                # SL
                if low <= sl:
                    exit_price = sl
                    exit_time = r['candle_time_bkk']
                    exit_reason = 'SL'
                    break
                # TSL
                if low <= tsl:
                    exit_price = tsl
                    exit_time = r['candle_time_bkk']
                    exit_reason = 'TSL'
                    break
            else:
                if low < lowest:
                    lowest = low
                    tsl = lowest + tsl_distance
                if low <= tp:
                    exit_price = tp
                    exit_time = r['candle_time_bkk']
                    exit_reason = 'TP'
                    break
                if high >= sl:
                    exit_price = sl
                    exit_time = r['candle_time_bkk']
                    exit_reason = 'SL'
                    break
                if high >= tsl:
                    exit_price = tsl
                    exit_time = r['candle_time_bkk']
                    exit_reason = 'TSL'
                    break

        if exit_price is None:
            last = candles_df.iloc[-1]
            exit_price = last['close']
            exit_time = last['candle_time_bkk']
            exit_reason = 'EndOfHistory'

        # compute P&L and capital change
        if side == "LONG":
            pnl_price = exit_price - entry_price
            pnl_pct = pnl_price / entry_price
        else:
            pnl_price = entry_price - exit_price
            pnl_pct = pnl_price / entry_price

        pnl_amount = capital * risk_factor * pnl_pct
        capital_before = capital
        capital_after = capital + pnl_amount
        capital = capital_after

        hist_rec = {
            "cross_index": entry_idx,
            "cross_type": cross['type'],
            "entry_time": entry_time,
            "entry_price": round(entry_price,2),
            "side": side,
            "SL": round(sl,2),
            "TP": round(tp,2),
            "TSL_at_entry": round(entry_price - tsl_distance if side=="LONG" else entry_price + tsl_distance,2),
            "exit_time": exit_time,
            "exit_price": round(exit_price,2),
            "exit_reason": exit_reason,
            "pnl_price": round(pnl_price,2),
            "pnl_pct": round(pnl_pct*100,4),
            "pnl_amount": round(pnl_amount,2),
            "capital_before": round(capital_before,2),
            "capital_after": round(capital_after,2)
        }
        hist_trades.append(hist_rec)

    return pd.DataFrame(hist_trades)

hist_trades_df = backtest_historical_with_tsl(candles, history_crosses, TSL_DISTANCE, capital_start, risk_factor)
# convert times to strings for safer Excel writing
if not hist_trades_df.empty:
    if 'entry_time' in hist_trades_df.columns:
        hist_trades_df['entry_time'] = hist_trades_df['entry_time'].astype(str)
    if 'exit_time' in hist_trades_df.columns:
        hist_trades_df['exit_time'] = hist_trades_df['exit_time'].astype(str)

swings_for_file = hist_trades_df.copy()

# ---------- Excel setup for this run ----------
run_tag = datetime.now(TZ_BKK).strftime("%Y%m%d_%H%M%S")
sheet_candles, sheet_swings, sheet_trades = ensure_excel_run_sheets(EXCEL_FILE, run_tag)

# Write initial historical data
write_df_safe(EXCEL_FILE, sheet_candles, candles)
write_df_safe(EXCEL_FILE, sheet_swings, swings_for_file)
write_df_safe(EXCEL_FILE, sheet_trades, pd.DataFrame(trades))  # empty now

print(f"Historical backtest swings_for_file: {swings_for_file}")

print(f"Historical backtest saved to sheet: {sheet_swings}. Starting live feed...")
print(f"Initial capital: {capital_start:.2f}")

# ---------- Live websocket logic ----------
def on_message(ws, message):
    global candles, history_crosses, trades, pending_order, capital, swings_for_file
    try:
        msg = json.loads(message)
        k = msg.get('k', {})
        if not k:
            return
        # Only process closed candles
        if not k.get('x', False):
            return

        t_open_ms = int(k['t'])
        t_open_dt = to_bkk(t_open_ms)
        o = round2(k['o']); h = round2(k['h']); l = round2(k['l']); c = round2(k['c']); v = round2(k['v'])

        # Append candle
        new_row = {
            "open": o, "high": h, "low": l, "close": c, "volume": v, "candle_time_bkk": t_open_dt
        }
        candles.loc[len(candles)] = new_row
        candles = calculate_emas(candles)

        last = candles.iloc[-1]
        prev = candles.iloc[-2] if len(candles) >= 2 else None

        # Print every bar close with BKK time and EMAs (safe formatting)
        time_str = last['candle_time_bkk'].strftime("%Y-%m-%d %H:%M:%S")
        ema_fast = last['EMA_fast'] if not pd.isna(last['EMA_fast']) else None
        ema_slow = last['EMA_slow'] if not pd.isna(last['EMA_slow']) else None
        close_str = f"{last['close']:.2f}"
        ema_fast_str = f"{ema_fast:.2f}" if ema_fast is not None else "NA"
        ema_slow_str = f"{ema_slow:.2f}" if ema_slow is not None else "NA"
        print(f"[{time_str}] Close: {close_str} EMA{EMA_FAST}: {ema_fast_str} EMA{EMA_SLOW}: {ema_slow_str}")

        # 1) If we have a pending_order placed at previous cross -> execute at this candle's OPEN
        if pending_order is not None:
            exec_price = o
            side = pending_order['side']
            placed_price = pending_order['placed_price']
            placed_time = pending_order['placed_time']
            placed_index = pending_order['placed_index']

            # initialize trade record values (TSL/SL/TP)
            if side == 'LONG':
                entry_price = exec_price
                highest = entry_price
                tsl = highest - TSL_DISTANCE
                sl = entry_price - TSL_DISTANCE
                tp = entry_price * (1 + tp_factor)
            else:  # SHORT
                entry_price = exec_price
                lowest = entry_price
                tsl = lowest + TSL_DISTANCE
                sl = entry_price + TSL_DISTANCE
                tp = entry_price * (1 - tp_factor)

            trade = {
                "cross_index": placed_index,
                "placed_time": str(placed_time),
                "placed_price": round2(placed_price),
                "side": side,
                "entry_time": time_str,
                "entry_price": round2(entry_price),
                "SL": round2(sl),
                "TP": round2(tp),
                "TSL": round2(tsl),
                "highest": round2(highest) if side=="LONG" else None,
                "lowest": round2(lowest) if side=="SHORT" else None,
                "status": "OPEN",
                "exit_time": "",
                "exit_price": "",
                "exit_reason": "",
                "pnl_amount": "",
                "capital_before": round2(capital),
                "capital_after": round2(capital)
            }
            trades.append(trade)
            print(f"\n[ENTRY CONFIRMED] Side: {side} | Placed at {placed_time} price {placed_price:.2f} | Executed at OPEN {entry_price:.2f}")
            print(f"    SL: {trade['SL']:.2f}  TP: {trade['TP']:.2f}  initial TSL: {trade['TSL']:.2f}\n")

            pending_order = None

        # 2) Detect cross on this closed candle -> place pending_order to execute at next open
        cross_event = None
        if prev is not None:
            if (prev['EMA_fast'] <= prev['EMA_slow']) and (last['EMA_fast'] > last['EMA_slow']):
                cross_event = {"index": len(candles)-1, "type":"GOLDEN", "time_bkk": last['candle_time_bkk'], "price": last['close']}
            elif (prev['EMA_fast'] >= prev['EMA_slow']) and (last['EMA_fast'] < last['EMA_slow']):
                cross_event = {"index": len(candles)-1, "type":"DEAD", "time_bkk": last['candle_time_bkk'], "price": last['close']}

            if cross_event:
                swings_for_file = swings_for_file  # leave as-is (we already saved historical swings)
                placed_side = "LONG" if cross_event['type']=="GOLDEN" else "SHORT"
                pending_order = {
                    "side": placed_side,
                    "placed_price": cross_event['price'],
                    "placed_time": cross_event['time_bkk'],
                    "placed_index": cross_event['index']
                }
                print(f"[SWING] {cross_event['type']} at {cross_event['time_bkk']} price {cross_event['price']:.2f}")
                print(f"[ORDER PLACED] Side: {placed_side}, placed at cross-candle CLOSE {cross_event['price']:.2f} -> will execute next candle OPEN\n")

        # 3) Manage open trades: update TSL only when price moves favorably, check TP/SL/TSL hits
        # iterate over active trades
        for t in list(trades):
            if t['status'] != "OPEN":
                continue
            side = t['side']
            entry_p = float(t['entry_price'])

            if side == "LONG":
                # update highest price and tsl if new high made in this candle
                cur_high = last['high']
                if cur_high > (t.get('highest') or entry_p):
                    t['highest'] = round2(cur_high)
                    t['TSL'] = round2(t['highest'] - TSL_DISTANCE)

                # Check TP hit (use candle high)
                if last['high'] >= float(t['TP']):
                    pnl_price = float(t['TP']) - entry_p
                    pnl_pct = pnl_price / entry_p
                    pnl_amount = capital * risk_factor * pnl_pct
                    cap_before = capital
                    capital += pnl_amount
                    t.update({
                        "status":"CLOSED",
                        "exit_time": time_str,
                        "exit_price": round2(float(t['TP'])),
                        "exit_reason":"TP",
                        "pnl_amount": round2(pnl_amount),
                        "capital_before": round2(cap_before),
                        "capital_after": round2(capital)
                    })
                    print(f"[ORDER] LONG TP HIT at {t['TP']:.2f} P&L:+{pnl_amount:.2f} NewCapital:{capital:.2f}")

                # Check SL or TSL hit (use candle low)
                elif last['low'] <= float(t['TSL']) or last['low'] <= float(t['SL']):
                    exit_price = float(t['TSL']) if last['low'] <= float(t['TSL']) else float(t['SL'])
                    pnl_price = exit_price - entry_p
                    pnl_pct = pnl_price / entry_p
                    pnl_amount = capital * risk_factor * pnl_pct
                    cap_before = capital
                    capital += pnl_amount
                    t.update({
                        "status":"CLOSED",
                        "exit_time": time_str,
                        "exit_price": round2(exit_price),
                        "exit_reason":"TSL" if last['low'] <= float(t['TSL']) else "SL",
                        "pnl_amount": round2(pnl_amount),
                        "capital_before": round2(cap_before),
                        "capital_after": round2(capital)
                    })
                    print(f"[ORDER] LONG EXIT at {exit_price:.2f} due to {'TSL' if last['low'] <= float(t['TSL']) else 'SL'}, P&L:{pnl_amount:.2f} NewCapital:{capital:.2f}")

            else:  # SHORT
                cur_low = last['low']
                if cur_low < (t.get('lowest') or entry_p):
                    t['lowest'] = round2(cur_low)
                    t['TSL'] = round2(t['lowest'] + TSL_DISTANCE)

                # TP check
                if last['low'] <= float(t['TP']):
                    pnl_price = entry_p - float(t['TP'])
                    pnl_pct = pnl_price / entry_p
                    pnl_amount = capital * risk_factor * pnl_pct
                    cap_before = capital
                    capital += pnl_amount
                    t.update({
                        "status":"CLOSED",
                        "exit_time": time_str,
                        "exit_price": round2(float(t['TP'])),
                        "exit_reason":"TP",
                        "pnl_amount": round2(pnl_amount),
                        "capital_before": round2(cap_before),
                        "capital_after": round2(capital)
                    })
                    print(f"[ORDER] SHORT TP HIT at {t['TP']:.2f} P&L:+{pnl_amount:.2f} NewCapital:{capital:.2f}")

                # SL or TSL hit (candle high)
                elif last['high'] >= float(t['TSL']) or last['high'] >= float(t['SL']):
                    exit_price = float(t['TSL']) if last['high'] >= float(t['TSL']) else float(t['SL'])
                    pnl_price = entry_p - exit_price
                    pnl_pct = pnl_price / entry_p
                    pnl_amount = capital * risk_factor * pnl_pct
                    cap_before = capital
                    capital += pnl_amount
                    t.update({
                        "status":"CLOSED",
                        "exit_time": time_str,
                        "exit_price": round2(exit_price),
                        "exit_reason":"TSL" if last['high'] >= float(t['TSL']) else "SL",
                        "pnl_amount": round2(pnl_amount),
                        "capital_before": round2(cap_before),
                        "capital_after": round2(capital)
                    })
                    print(f"[ORDER] SHORT EXIT at {exit_price:.2f} due to {'TSL' if last['high'] >= float(t['TSL']) else 'SL'}, P&L:{pnl_amount:.2f} NewCapital:{capital:.2f}")

        # ---------- Save to Excel ----------
        try:
            write_df_safe(EXCEL_FILE, sheet_candles, candles)
            # Write historical swings/backtest (unchanged) to swings sheet
            write_df_safe(EXCEL_FILE, sheet_swings, swings_for_file if 'swings_for_file' in globals() else pd.DataFrame())
            # Write trades (historical backtest rows not appended to trades list; trades are live executed entries)
            trades_df = pd.DataFrame(trades)
            write_df_safe(EXCEL_FILE, sheet_trades, trades_df)
        except Exception as e:
            print("Excel write warning:", e)

    except Exception as exc:
        print("Error in on_message:", exc)
        traceback.print_exc()

def on_open(ws):
    print("WebSocket opened.")

def on_error(ws, error):
    print("WebSocket error:", error)

def on_close(ws, close_status_code, close_msg):
    print("WebSocket closed.", close_status_code, close_msg)

# ---------- Start WebSocket ----------
ws_url = f"wss://stream.binance.com:9443/ws/{SYMBOL.lower()}@kline_{INTERVAL}"
print(f"Starting live EMA {EMA_FAST}-{EMA_SLOW} strategy for {SYMBOL} with capital {capital_start:.2f}")
ws = websocket.WebSocketApp(ws_url, on_open=on_open, on_message=on_message, on_error=on_error, on_close=on_close)

try:
    ws.run_forever()
except KeyboardInterrupt:
    print("Stopped by user.")
