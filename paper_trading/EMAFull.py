# ema_trailing_live_final.py
import websocket
import json
import pandas as pd
import requests
from datetime import datetime
import pytz
import os
from openpyxl import load_workbook
import time

# ---------------- USER SETTINGS ----------------
symbol = "BTCUSDT"
interval = "1m"
history_limit = 500

# Strategy parameters
# === USER SETTINGS (percent based) ===
stop_loss_pct = 1               # % stop loss (NOT 0.02)
target_pct = 1.5                  # % take profit
risk_per_trade_pct = 1          # % of capital used for P&L scaling
TSL_DISTANCE = 200.0            # trailing stop distance in points
capital_start = 100000.0

# --- EMA SETTINGS ---
EMA_FAST = 9
EMA_SLOW = 21

# === INTERNAL CONVERSION ===
sl_factor   = stop_loss_pct / 100.0
tp_factor   = target_pct / 100.0
risk_factor = risk_per_trade_pct / 100.0

ts = datetime.now().strftime("%Y%m%d_%H%M%S")
excel_file = f"excel_data/{symbol}_{ts}.xlsx"
tz_bkk = pytz.timezone("Asia/Bangkok")
# ------------------------------------------------

# In-memory containers
candles = pd.DataFrame(columns=['open','high','low','close','volume','candle_time_bkk','EMA9','EMA21'])
swings = []     # historical + live cross events (we will write backtest results here)
trades = []     # executed trades (history + live)
pending_order = None  # {'side':'LONG'/'SHORT', 'placed_time':..., 'placed_price':..., 'placed_index':...}
capital = capital_start

# ---------- Helpers ----------
def round2(v):
    return round(float(v), 2)

def to_bkk(ts_ms):
    return datetime.fromtimestamp(ts_ms/1000, pytz.utc).astimezone(tz_bkk)

def to_bkk_str(ts_ms):
    return to_bkk(ts_ms).strftime("%Y-%m-%d %H:%M:%S")

def binance_klines(symbol, interval, limit=500):
    url = "https://api.binance.com/api/v3/klines"
    params = {"symbol": symbol, "interval": interval, "limit": limit}
    r = requests.get(url, params=params, timeout=10)
    r.raise_for_status()
    return r.json()

def calculate_emas(df):
    if len(df) >= 1:
        df['EMA9'] = df['close'].ewm(span=9, adjust=False).mean()
        df['EMA21'] = df['close'].ewm(span=21, adjust=False).mean()
    return df

def detect_crosses(df):
    crosses = []
    if 'EMA9' not in df or 'EMA21' not in df:
        return crosses
    for i in range(1, len(df)):
        a = df.iloc[i-1]
        b = df.iloc[i]
        if pd.isna(a['EMA9']) or pd.isna(a['EMA21']) or pd.isna(b['EMA9']) or pd.isna(b['EMA21']):
            continue
        if (a['EMA9'] <= a['EMA21']) and (b['EMA9'] > b['EMA21']):
            crosses.append({"index": i, "type": "GOLDEN", "time_bkk": b['candle_time_bkk'], "price": b['close']})
        if (a['EMA9'] >= a['EMA21']) and (b['EMA9'] < b['EMA21']):
            crosses.append({"index": i, "type": "DEAD", "time_bkk": b['candle_time_bkk'], "price": b['close']})
    return crosses

def ensure_excel_run_sheets(base_file, run_tag):
    """Create three timestamped sheets for this run. Return names."""
    sheet_candles = f"Candles"
    sheet_swings = f"Swings"
    sheet_trades = f"Trades"
    if not os.path.exists(base_file):
        writer = pd.ExcelWriter(base_file, engine='openpyxl', mode='w')
        pd.DataFrame().to_excel(writer, sheet_name=sheet_candles, index=False)
        pd.DataFrame().to_excel(writer, sheet_name=sheet_swings, index=False)
        pd.DataFrame().to_excel(writer, sheet_name=sheet_trades, index=False)
        writer.close()
    else:
        wb = load_workbook(base_file)
        def unique(name):
            if name not in wb.sheetnames:
                return name
            n = 1
            while f"{name}_{n}" in wb.sheetnames:
                n += 1
            return f"{name}_{n}"
        sheet_candles = unique(sheet_candles)
        sheet_swings = unique(sheet_swings)
        sheet_trades = unique(sheet_trades)
        writer = pd.ExcelWriter(base_file, engine='openpyxl', mode='a')
        pd.DataFrame().to_excel(writer, sheet_name=sheet_candles, index=False)
        pd.DataFrame().to_excel(writer, sheet_name=sheet_swings, index=False)
        pd.DataFrame().to_excel(writer, sheet_name=sheet_trades, index=False)
        writer.close()
    return sheet_candles, sheet_swings, sheet_trades

def write_df_safe(base_file, sheet_name, df):
    """Write DataFrame to Excel sheet after making datetimes Excel-safe and rounding values."""
    df_copy = df.copy()
    # Convert datetime columns to naive (Excel-friendly)
    for col in df_copy.columns:
        if 'time' in col and pd.api.types.is_datetime64_any_dtype(df_copy[col]):
            df_copy[col] = df_copy[col].dt.tz_localize(None)
    # Round numeric columns to 2 decimals for neatness
    for c in df_copy.columns:
        if pd.api.types.is_float_dtype(df_copy[c]):
            df_copy[c] = df_copy[c].round(2)
    with pd.ExcelWriter(base_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_copy.to_excel(writer, sheet_name=sheet_name, index=False)

# ---------- Historical preload & backtest ----------
print("Fetching historical candles...")
klines = binance_klines(symbol, interval, history_limit)
rows = []
for k in klines:
    rows.append({
        "open": round2(k[1]),
        "high": round2(k[2]),
        "low": round2(k[3]),
        "close": round2(k[4]),
        "volume": round2(k[5]),
        "candle_time_bkk": to_bkk(k[0])
    })
candles = pd.DataFrame(rows)
candles = calculate_emas(candles)
print(f"Loaded {len(candles)} historical candles.")

# detect historical crosses
history_crosses = detect_crosses(candles)
print(f"Detected {len(history_crosses)} historical crosses.")

# Historical backtest using Option A TSL behavior (point-based)
def backtest_historical_with_tsl(candles_df, crosses, tsl_distance, capital_start, risk_per_trade_pct):
    capital = capital_start
    hist_trades = []
    for i, cross in enumerate(crosses):
        entry_idx = cross['index']
        entry_price = candles_df.iloc[entry_idx]['close']
        entry_time = candles_df.iloc[entry_idx]['candle_time_bkk']
        side = "LONG" if cross['type'] == "GOLDEN" else "SHORT"

        # Initialize TSL / trackers
        if side == "LONG":
            tsl = entry_price - tsl_distance
            highest = entry_price
            sl = entry_price - tsl_distance  # you requested point-based SL identical to TSL distance; keep same
            tp = entry_price * (1 + target_pct)
        else:
            tsl = entry_price + tsl_distance
            lowest = entry_price
            sl = entry_price + tsl_distance
            tp = entry_price * (1 - target_pct)

        exit_price = None
        exit_time = None
        exit_reason = None

        # iterate forward to find exit
        for j in range(entry_idx + 1, len(candles_df)):
            row = candles_df.iloc[j]
            high = row['high']
            low = row['low']
            close = row['close']

            if side == "LONG":
                # update highest and tsl if price made new high
                if high > highest:
                    highest = high
                    tsl = highest - tsl_distance
                # check TP
                if high >= tp:
                    exit_price = tp
                    exit_time = row['candle_time_bkk']
                    exit_reason = 'TP'
                    break
                # check SL or TSL hit
                if low <= sl:
                    exit_price = sl
                    exit_time = row['candle_time_bkk']
                    exit_reason = 'SL'
                    break
                if low <= tsl:
                    exit_price = tsl
                    exit_time = row['candle_time_bkk']
                    exit_reason = 'TSL'
                    break
            else:  # SHORT
                if low < lowest:
                    lowest = low
                    tsl = lowest + tsl_distance
                if low <= tp:
                    exit_price = tp
                    exit_time = row['candle_time_bkk']
                    exit_reason = 'TP'
                    break
                if high >= sl:
                    exit_price = sl
                    exit_time = row['candle_time_bkk']
                    exit_reason = 'SL'
                    break
                if high >= tsl:
                    exit_price = tsl
                    exit_time = row['candle_time_bkk']
                    exit_reason = 'TSL'
                    break

        # if no exit found, close at last candle close (EndOfHistory)
        if exit_price is None:
            last = candles_df.iloc[-1]
            exit_price = last['close']
            exit_time = last['candle_time_bkk']
            exit_reason = 'EndOfHistory'

        # compute pnl (price-based) and convert to capital change using risk_per_trade_pct
        if side == "LONG":
            pnl_price = exit_price - entry_price
            pnl_pct = pnl_price / entry_price
        else:
            pnl_price = entry_price - exit_price
            pnl_pct = pnl_price / entry_price

        pnl_amount = capital * risk_per_trade_pct * pnl_pct
        capital_before = capital
        capital_after = capital + pnl_amount
        capital = capital_after

        hist_trades.append({
            "cross_index": entry_idx,
            "cross_type": cross['type'],
            "placed_time": entry_time,
            "entry_time": entry_time,
            "entry_price": round(entry_price,2),
            "side": side,
            "SL": round(sl,2),
            "TP": round(tp,2),
            "TSL_at_entry": round(entry_price - tsl_distance if side=="LONG" else entry_price + tsl_distance,2),
            "exit_time": exit_time,
            "exit_price": round(exit_price,2),
            "exit_reason": exit_reason,
            "pnl_price": round(pnl_price,2),
            "pnl_pct": round(pnl_pct*100,4),
            "pnl_amount": round(pnl_amount,2),
            "capital_before": round(capital_before,2),
            "capital_after": round(capital_after,2)
        })

    return pd.DataFrame(hist_trades)

# run historical backtest and store in swings (for initial save)
hist_trades_df = backtest_historical_with_tsl(candles, history_crosses, TSL_DISTANCE, capital_start, risk_per_trade_pct)
# add to swings list as records (so live file will contain history)
# convert times to strings for safety when writing
if not hist_trades_df.empty:
    hist_trades_df['placed_time'] = hist_trades_df['placed_time'].astype(str)
    hist_trades_df['entry_time'] = hist_trades_df['entry_time'].astype(str)
    hist_trades_df['exit_time'] = hist_trades_df['exit_time'].astype(str)
swings_for_file = hist_trades_df.copy()

# ---------- Excel setup for this run ----------
run_tag = datetime.now(tz_bkk).strftime("%Y%m%d_%H%M%S")
sheet_candles, sheet_swings, sheet_trades = ensure_excel_run_sheets(excel_file, run_tag)

# Write initial historical candles and swings/trades to Excel
write_df_safe(excel_file, sheet_candles, candles)
write_df_safe(excel_file, sheet_swings, swings_for_file)
# no executed live trades yet; write empty trades sheet
write_df_safe(excel_file, sheet_trades, pd.DataFrame(trades))

print(f"Historical backtest saved to sheet {sheet_swings}. Starting live websocket...")

# ---------- Live websocket with pending-order -> execute at next candle OPEN ----------
def on_message(ws, message):
    global candles, swings, trades, pending_order, capital
    try:
        msg = json.loads(message)
        k = msg.get('k', {})
        if not k:
            return

        # Only process when candle closed
        if not k.get('x', False):
            return

        # closed candle data
        t_open_ms = int(k['t'])
        t_open = to_bkk(t_open_ms)
        o = round2(k['o'])
        h = round2(k['h'])
        l = round2(k['l'])
        c = round2(k['c'])
        v = round2(k['v'])

        # Append candle
        new_row = {
            "open": o, "high": h, "low": l, "close": c, "volume": v, "candle_time_bkk": t_open
        }
        candles.loc[len(candles)] = new_row
        candles = calculate_emas(candles)

        last = candles.iloc[-1]
        prev = candles.iloc[-2] if len(candles) >= 2 else None

        # Print every bar close with BKK time and EMAs
        # convert timezone-aware to string for printing
        time_str = last['candle_time_bkk'].strftime("%Y-%m-%d %H:%M:%S")
        ema9 = last['EMA9'] if not pd.isna(last['EMA9']) else "NA"
        ema21 = last['EMA21'] if not pd.isna(last['EMA21']) else "NA"
        ema9_str = f"{ema9:.2f}" if ema9 is not None else "NA"
        ema21_str = f"{ema21:.2f}" if ema21 is not None else "NA"
        close_str = f"{c:.2f}" if c is not None else "NA"

        print(f"[{time_str}] Close: {close_str} EMA9: {ema9_str} EMA21: {ema21_str}")

        # ---------- First: if we have a pending_order from previous cross, execute it at THIS candle's OPEN ----------
        # pending_order structure: {'side': 'LONG'/'SHORT', 'placed_price':..., 'placed_time':..., 'placed_index':...}
        if pending_order is not None:
            exec_price = o  # execute at this candle's open
            side = pending_order['side']
            placed_price = pending_order['placed_price']
            placed_time = pending_order['placed_time']
            placed_index = pending_order['placed_index']

            # Create the executed trade record with TSL initialised per Option A
            if side == 'LONG':
                entry_price = exec_price
                highest = entry_price
                tsl = highest - TSL_DISTANCE
                sl = entry_price - TSL_DISTANCE
                tp = entry_price * (1 + target_pct)
            else:
                entry_price = exec_price
                lowest = entry_price
                tsl = lowest + TSL_DISTANCE
                sl = entry_price + TSL_DISTANCE
                tp = entry_price * (1 - target_pct)

            trade = {
                "cross_index": placed_index,
                "placed_time": str(placed_time),
                "placed_price": round2(placed_price),
                "side": side,
                "entry_time": time_str,
                "entry_price": round2(entry_price),
                "SL": round2(sl),
                "TP": round2(tp),
                "TSL": round2(tsl),
                "highest": round2(highest) if side=="LONG" else None,
                "lowest": round2(lowest) if side=="SHORT" else None,
                "status": "OPEN",
                "exit_time": "",
                "exit_price": "",
                "exit_reason": "",
                "pnl_amount": "",
                "capital_before": round2(capital),
                "capital_after": round2(capital)
            }
            trades.append(trade)
            print(f"\n[ENTRY CONFIRMED] Side: {side} | Placed at {placed_time} price {placed_price:.2f} | Executed at OPEN {entry_price:.2f}")
            print(f"   SL: {trade['SL']:.2f}  TP: {trade['TP']:.2f}  initial TSL: {trade['TSL']:.2f}\n")

            pending_order = None  # clear pending

        # ---------- Then: detect cross on this closed candle and PLACE pending order (to execute at next open) ----------
        cross_event = None
        if prev is not None:
            if (prev['EMA9'] <= prev['EMA21']) and (last['EMA9'] > last['EMA21']):
                cross_event = {"index": len(candles)-1, "type": "GOLDEN", "time_bkk": last['candle_time_bkk'], "price": last['close']}
            elif (prev['EMA9'] >= prev['EMA21']) and (last['EMA9'] < last['EMA21']):
                cross_event = {"index": len(candles)-1, "type": "DEAD", "time_bkk": last['candle_time_bkk'], "price": last['close']}

            if cross_event:
                # store cross in swings (for logging/history)
                swings.append(cross_event)
                placed_side = "LONG" if cross_event['type']=="GOLDEN" else "SHORT"
                pending_order = {
                    "side": placed_side,
                    "placed_price": cross_event['price'],
                    "placed_time": cross_event['time_bkk'],
                    "placed_index": cross_event['index']
                }
                print(f"[SWING] {cross_event['type']} at {cross_event['time_bkk']} price {cross_event['price']:.2f}")
                print(f"[ORDER PLACED] Side: {placed_side}, placed at cross-candle CLOSE {cross_event['price']:.2f} -> will execute next candle OPEN\n")

        # ---------- Manage open trades: update TSL and check for TP/SL/TSL hits ----------
        # ITERATE over a copy to avoid modifying list while iterating
        for t in list(trades):
            if t['status'] != "OPEN":
                continue
            side = t['side']
            entry_p = float(t['entry_price'])
            # LONG trade management
            if side == "LONG":
                # update highest price and TSL if new high made in this candle
                if last['high'] > (t.get('highest') or entry_p):
                    t['highest'] = round2(last['high'])
                    t['TSL'] = round2(t['highest'] - TSL_DISTANCE)
                # check TP
                if last['high'] >= float(t['TP']):
                    # hit TP
                    pnl_price = float(t['TP']) - entry_p
                    pnl_pct = pnl_price / entry_p
                    pnl_amount = capital * risk_per_trade_pct * pnl_pct
                    cap_before = capital
                    capital += pnl_amount
                    t.update({
                        "status":"CLOSED",
                        "exit_time": time_str,
                        "exit_price": round2(float(t['TP'])),
                        "exit_reason":"TP",
                        "pnl_amount": round2(pnl_amount),
                        "capital_before": round2(cap_before),
                        "capital_after": round2(capital)
                    })
                    print(f"[ORDER] LONG TP HIT at {t['TP']:.2f} P&L:+{pnl_amount:.2f} NewCapital:{capital:.2f}")
                # check SL or TSL hit (TSL uses last['low'] to see if price fell below)
                elif last['low'] <= float(t['TSL']) or last['low'] <= float(t['SL']):
                    # Exited by TSL or SL. Use TSL if both triggered by gap? We'll prioritize TSL if low <= TSL.
                    exit_price = float(t['TSL']) if last['low'] <= float(t['TSL']) else float(t['SL'])
                    pnl_price = exit_price - entry_p
                    pnl_pct = pnl_price / entry_p
                    pnl_amount = capital * risk_per_trade_pct * pnl_pct
                    cap_before = capital
                    capital += pnl_amount
                    t.update({
                        "status":"CLOSED",
                        "exit_time": time_str,
                        "exit_price": round2(exit_price),
                        "exit_reason":"TSL" if last['low'] <= float(t['TSL']) else "SL",
                        "pnl_amount": round2(pnl_amount),
                        "capital_before": round2(cap_before),
                        "capital_after": round2(capital)
                    })
                    print(f"[ORDER] LONG EXIT at {exit_price:.2f} due to {'TSL' if last['low'] <= float(t['TSL']) else 'SL'}, P&L:{pnl_amount:.2f} NewCapital:{capital:.2f}")

            else:  # SHORT
                # update lowest price and TSL if new low made in this candle
                if last['low'] < (t.get('lowest') or entry_p):
                    t['lowest'] = round2(last['low'])
                    t['TSL'] = round2(t['lowest'] + TSL_DISTANCE)
                # TP check (price falls)
                if last['low'] <= float(t['TP']):
                    pnl_price = entry_p - float(t['TP'])
                    pnl_pct = pnl_price / entry_p
                    pnl_amount = capital * risk_per_trade_pct * pnl_pct
                    cap_before = capital
                    capital += pnl_amount
                    t.update({
                        "status":"CLOSED",
                        "exit_time": time_str,
                        "exit_price": round2(float(t['TP'])),
                        "exit_reason":"TP",
                        "pnl_amount": round2(pnl_amount),
                        "capital_before": round2(cap_before),
                        "capital_after": round2(capital)
                    })
                    print(f"[ORDER] SHORT TP HIT at {t['TP']:.2f} P&L:+{pnl_amount:.2f} NewCapital:{capital:.2f}")
                # SL or TSL hit
                elif last['high'] >= float(t['TSL']) or last['high'] >= float(t['SL']):
                    exit_price = float(t['TSL']) if last['high'] >= float(t['TSL']) else float(t['SL'])
                    pnl_price = entry_p - exit_price
                    pnl_pct = pnl_price / entry_p
                    pnl_amount = capital * risk_per_trade_pct * pnl_pct
                    cap_before = capital
                    capital += pnl_amount
                    t.update({
                        "status":"CLOSED",
                        "exit_time": time_str,
                        "exit_price": round2(exit_price),
                        "exit_reason":"TSL" if last['high'] >= float(t['TSL']) else "SL",
                        "pnl_amount": round2(pnl_amount),
                        "capital_before": round2(cap_before),
                        "capital_after": round2(capital)
                    })
                    print(f"[ORDER] SHORT EXIT at {exit_price:.2f} due to {'TSL' if last['high'] >= float(t['TSL']) else 'SL'}, P&L:{pnl_amount:.2f} NewCapital:{capital:.2f}")

        # ---------- Save everything to Excel (candles, swings history+backtest, trades) ----------
        try:
            # Save candles
            write_df_safe(excel_file, sheet_candles, candles)
            # Save swings: combine historical backtest results (hist_trades_df) with live swings (as simple records)
            # We will write the historical backtest (already saved at start) and also append live swings summary
            live_swings_df = pd.DataFrame(swings)
            # if there are no columns (empty), keep as empty DataFrame
            write_df_safe(excel_file, sheet_swings, swings_for_file if 'swings_for_file' in globals() else pd.DataFrame())
            # Save trades (history + live)
            trades_df = pd.DataFrame(trades)
            write_df_safe(excel_file, sheet_trades, trades_df)
        except Exception as e:
            print("Excel write warning:", e)

    except Exception as exc:
        print("Error in on_message:", exc)

def on_open(ws):
    print("WebSocket opened.")

def on_error(ws, error):
    print("WebSocket error:", error)

def on_close(ws, close_status_code, close_msg):
    print("WebSocket closed.", close_status_code, close_msg)

# ---------- Start websocket ----------
ws_url = f"wss://stream.binance.com:9443/ws/{symbol.lower()}@kline_{interval}"
print(f"Starting live EMA 9-21 with TSL_DISTANCE={TSL_DISTANCE}, capital={capital_start}")
ws = websocket.WebSocketApp(ws_url, on_open=on_open, on_message=on_message, on_error=on_error, on_close=on_close)

try:
    ws.run_forever()
except KeyboardInterrupt:
    print("Stopped by user.")
