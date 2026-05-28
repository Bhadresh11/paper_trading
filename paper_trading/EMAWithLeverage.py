# ema_trailing_live_with_leverage.py
import websocket
import json
import pandas as pd
import requests
from datetime import datetime
import pytz
import os
from openpyxl import load_workbook
import traceback

# ===========================
# USER CONFIGURATION PANEL
# ===========================
SYMBOL = "BTCUSDT"
INTERVAL = "1m"
HISTORY_LIMIT = 500

# Capital (USD)
capital_start = 100000.0  # user said 100000 USD

# EMA settings
EMA_FAST = 5
EMA_SLOW = 20

# Percent settings (user-friendly 0-100)
stop_loss_pct = 2        # percent (not decimal)
target_pct = 1           # percent TP
risk_per_trade_pct = 1   # percent (if you want to use percent-based risk for alternate sizing)

# Trailing stop distance (points)
TSL_DISTANCE = 100.0     # points (example)

# Lot / leverage settings (your choices)
LOT_MODE = "A"           # A: 1 lot = 0.001 BTC
LOTS_PER_TRADE = 1000    # default lots per trade -> 1000 * 0.001 = 1 BTC position
LEVERAGE = 100           # 100x

# Excel output file
# EXCEL_FILE = "EMA_TSL_leverage_live.xlsx"
ts = datetime.now().strftime("%m%d_%H%M%S")
EXCEL_FILE = f"excel_data/{EMA_SLOW}_{EMA_FAST}_leverage_{ts}.xlsx"
# Timezone
TZ_BKK = pytz.timezone("Asia/Bangkok")
# ===========================
# Internal derived factors
sl_factor = stop_loss_pct / 100.0
tp_factor = target_pct / 100.0
risk_factor = risk_per_trade_pct / 100.0

# Lot conversion: interpret LOT_MODE A = 1 lot = 0.001 BTC
if LOT_MODE == "A":
    BTC_PER_LOT = 0.001
else:
    BTC_PER_LOT = 1.0  # B mode (not used since you chose A)

# ========== in-memory containers ==========
candles = pd.DataFrame(columns=[
    'open','high','low','close','volume','candle_time_bkk','EMA_fast','EMA_slow'
])
history_crosses = []
hist_trades_df = pd.DataFrame()
swings_for_file = pd.DataFrame()
trades = []            # list of executed trade dicts (live and closed)
pending_order = None   # store order placed at cross-candle close; executed next candle open
capital = capital_start

# ========== helpers ==========
def round2(v): return round(float(v), 2)

def to_bkk(ts_ms): return datetime.fromtimestamp(ts_ms/1000, tz=pytz.utc).astimezone(TZ_BKK)

def binance_klines(symbol, interval, limit=500):
    url = "https://api.binance.com/api/v3/klines"
    params = {"symbol": symbol, "interval": interval, "limit": limit}
    r = requests.get(url, params=params, timeout=10)
    r.raise_for_status()
    return r.json()

def calculate_emas(df):
    if len(df) >= 1:
        df['EMA_fast'] = df['close'].ewm(span=EMA_FAST, adjust=False).mean()
        df['EMA_slow'] = df['close'].ewm(span=EMA_SLOW, adjust=False).mean()
    return df

def detect_crosses(df):
    crosses = []
    if 'EMA_fast' not in df or 'EMA_slow' not in df:
        return crosses
    for i in range(1, len(df)):
        a = df.iloc[i-1]; b = df.iloc[i]
        if pd.isna(a['EMA_fast']) or pd.isna(a['EMA_slow']) or pd.isna(b['EMA_fast']) or pd.isna(b['EMA_slow']):
            continue
        if (a['EMA_fast'] <= a['EMA_slow']) and (b['EMA_fast'] > b['EMA_slow']):
            crosses.append({"index": i, "type": "GOLDEN", "time_bkk": b['candle_time_bkk'], "price": b['close']})
        if (a['EMA_fast'] >= a['EMA_slow']) and (b['EMA_fast'] < b['EMA_slow']):
            crosses.append({"index": i, "type": "DEAD", "time_bkk": b['candle_time_bkk'], "price": b['close']})
    return crosses

def ensure_excel_run_sheets(base_file, run_tag):
    sheet_candles = f"Candles"
    sheet_swings = f"Swings"
    sheet_trades = f"Trades"
    if not os.path.exists(base_file):
        writer = pd.ExcelWriter(base_file, engine='openpyxl', mode='w')
        pd.DataFrame().to_excel(writer, sheet_name=sheet_candles, index=False)
        pd.DataFrame().to_excel(writer, sheet_name=sheet_swings, index=False)
        pd.DataFrame().to_excel(writer, sheet_name=sheet_trades, index=False)
        writer.close()
    else:
        wb = load_workbook(base_file)
        def unique(name):
            if name not in wb.sheetnames: return name
            n = 1
            while f"{name}_{n}" in wb.sheetnames: n += 1
            return f"{name}_{n}"
        sheet_candles = unique(sheet_candles)
        sheet_swings = unique(sheet_swings)
        sheet_trades = unique(sheet_trades)
        writer = pd.ExcelWriter(base_file, engine='openpyxl', mode='a')
        pd.DataFrame().to_excel(writer, sheet_name=sheet_candles, index=False)
        pd.DataFrame().to_excel(writer, sheet_name=sheet_swings, index=False)
        pd.DataFrame().to_excel(writer, sheet_name=sheet_trades, index=False)
        writer.close()
    return sheet_candles, sheet_swings, sheet_trades

def write_df_safe(base_file, sheet_name, df):
    df_copy = df.copy()
    # convert timezone-aware datetimes to naive datetimes for Excel
    for col in df_copy.columns:
        if 'time' in col and col in df_copy and pd.api.types.is_datetime64_any_dtype(df_copy[col]):
            try:
                df_copy[col] = df_copy[col].dt.tz_convert(None).dt.tz_localize(None)
            except Exception:
                try:
                    df_copy[col] = df_copy[col].dt.tz_localize(None)
                except Exception:
                    pass
    # round floats
    for c in df_copy.columns:
        if pd.api.types.is_float_dtype(df_copy[c]) or pd.api.types.is_integer_dtype(df_copy[c]):
            try:
                df_copy[c] = df_copy[c].round(2)
            except Exception:
                pass
    with pd.ExcelWriter(base_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_copy.to_excel(writer, sheet_name=sheet_name, index=False)

# ========== Historical preload & backtest (position-based P&L) ==========
print("Fetching historical candles...")
klines = binance_klines(SYMBOL, INTERVAL, HISTORY_LIMIT)
rows = []
for k in klines:
    rows.append({
        "open": round2(k[1]),
        "high": round2(k[2]),
        "low": round2(k[3]),
        "close": round2(k[4]),
        "volume": round2(k[5]),
        "candle_time_bkk": to_bkk(k[0])
    })
candles = pd.DataFrame(rows)
candles = calculate_emas(candles)
print(f"Loaded {len(candles)} historical candles.")

history_crosses = detect_crosses(candles)
print(f"Detected {len(history_crosses)} historical cross events.")

def backtest_historical_position_pnl(candles_df, crosses, tsl_distance, lots_per_trade, btc_per_lot, leverage, capital_start):
    capital = capital_start
    records = []
    position_btc = lots_per_trade * btc_per_lot
    for cross in crosses:
        entry_idx = cross['index']
        entry_price = candles_df.iloc[entry_idx]['close']
        entry_time = candles_df.iloc[entry_idx]['candle_time_bkk']
        side = "LONG" if cross['type'] == "GOLDEN" else "SHORT"

        # initialize tsl and trackers
        if side == "LONG":
            highest = entry_price
            tsl = highest - tsl_distance
            sl = entry_price - tsl_distance
            tp = entry_price * (1 + tp_factor)
        else:
            lowest = entry_price
            tsl = lowest + tsl_distance
            sl = entry_price + tsl_distance
            tp = entry_price * (1 - tp_factor)

        exit_price = None; exit_time = None; exit_reason = None

        # simulate forward to find exit
        for j in range(entry_idx+1, len(candles_df)):
            r = candles_df.iloc[j]
            high = r['high']; low = r['low']; close = r['close']
            if side == "LONG":
                if high > highest:
                    highest = high; tsl = highest - tsl_distance
                if high >= tp:
                    exit_price = tp; exit_time = r['candle_time_bkk']; exit_reason='TP'; break
                if low <= sl:
                    exit_price = sl; exit_time = r['candle_time_bkk']; exit_reason='SL'; break
                if low <= tsl:
                    exit_price = tsl; exit_time = r['candle_time_bkk']; exit_reason='TSL'; break
            else:
                if low < lowest:
                    lowest = low; tsl = lowest + tsl_distance
                if low <= tp:
                    exit_price = tp; exit_time = r['candle_time_bkk']; exit_reason='TP'; break
                if high >= sl:
                    exit_price = sl; exit_time = r['candle_time_bkk']; exit_reason='SL'; break
                if high >= tsl:
                    exit_price = tsl; exit_time = r['candle_time_bkk']; exit_reason='TSL'; break

        if exit_price is None:
            last = candles_df.iloc[-1]
            exit_price = last['close']; exit_time = last['candle_time_bkk']; exit_reason='EndOfHistory'

        # position P&L (USD): price change * position BTC
        if side == "LONG":
            pnl_usd = (exit_price - entry_price) * position_btc
        else:
            pnl_usd = (entry_price - exit_price) * position_btc

        capital_before = capital
        capital_after = capital + pnl_usd
        capital = capital_after

        # margin used at entry = notional / leverage
        notional = entry_price * position_btc
        margin_used = notional / leverage

        records.append({
            "cross_index": entry_idx,
            "cross_type": cross['type'],
            "entry_time": str(entry_time),
            "entry_price": round(entry_price,2),
            "side": side,
            "position_btc": position_btc,
            "lots": lots_per_trade,
            "BTC_per_lot": btc_per_lot,
            "leverage": leverage,
            "notional_usd": round(notional,2),
            "margin_used": round(margin_used,2),
            "SL": round(sl,2),
            "TP": round(tp,2),
            "TSL_at_exit": round(exit_price,2),
            "exit_time": str(exit_time),
            "exit_price": round(exit_price,2),
            "exit_reason": exit_reason,
            "pnl_usd": round(pnl_usd,2),
            "capital_before": round(capital_before,2),
            "capital_after": round(capital_after,2)
        })

    return pd.DataFrame(records)

hist_trades_df = backtest_historical_position_pnl(
    candles, history_crosses, TSL_DISTANCE,
    LOTS_PER_TRADE, BTC_PER_LOT, LEVERAGE, capital_start
)

# save historical results (Swings) and candles
run_tag = datetime.now(TZ_BKK).strftime("%Y%m%d_%H%M%S")
sheet_candles, sheet_swings, sheet_trades = ensure_excel_run_sheets(EXCEL_FILE, run_tag)
write_df_safe(EXCEL_FILE, sheet_candles, candles)
write_df_safe(EXCEL_FILE, sheet_swings, hist_trades_df)
write_df_safe(EXCEL_FILE, sheet_trades, pd.DataFrame(trades))

print(f"Historical backtest hist_trades_df {hist_trades_df}")
print(f"Historical backtest saved to sheet {sheet_swings}. Starting live websocket...")
print(f"Initial capital: {capital_start:.2f} USD")
# ========== Live websocket (pending->execute at next open) ==========
def on_message(ws, message):
    global candles, pending_order, trades, capital, hist_trades_df
    try:
        msg = json.loads(message)
        k = msg.get('k', {})
        if not k:
            return
        # proceed only on closed candle
        if not k.get('x', False):
            return

        t_open_ms = int(k['t']); t_open = to_bkk(t_open_ms)
        o = round2(k['o']); h = round2(k['h']); l = round2(k['l']); c = round2(k['c']); v = round2(k['v'])
        candles.loc[len(candles)] = {"open":o,"high":h,"low":l,"close":c,"volume":v,"candle_time_bkk":t_open}
        candles = calculate_emas(candles)
        last = candles.iloc[-1]; prev = candles.iloc[-2] if len(candles)>=2 else None

        # print bar info
        time_str = last['candle_time_bkk'].strftime("%Y-%m-%d %H:%M:%S")
        ema_fast = last['EMA_fast'] if not pd.isna(last['EMA_fast']) else None
        ema_slow = last['EMA_slow'] if not pd.isna(last['EMA_slow']) else None
        ema_fast_str = f"{ema_fast:.2f}" if ema_fast is not None else "NA"
        ema_slow_str = f"{ema_slow:.2f}" if ema_slow is not None else "NA"
        print(f"[{time_str}] Close: {last['close']:.2f} EMA{EMA_FAST}: {ema_fast_str} EMA{EMA_SLOW}: {ema_slow_str}")

        # 1) execute pending order at this candle open (o)
        if pending_order is not None:
            exec_price = o
            side = pending_order['side']
            placed_price = pending_order['placed_price']
            placed_time = pending_order['placed_time']
            placed_index = pending_order['placed_index']

            # position
            position_btc = LOTS_PER_TRADE * BTC_PER_LOT
            notional = exec_price * position_btc
            margin_used = notional / LEVERAGE

            if side == "LONG":
                entry_price = exec_price
                highest = entry_price
                tsl = highest - TSL_DISTANCE
                sl = entry_price - TSL_DISTANCE
                tp = entry_price * (1 + tp_factor)
            else:
                entry_price = exec_price
                lowest = entry_price
                tsl = lowest + TSL_DISTANCE
                sl = entry_price + TSL_DISTANCE
                tp = entry_price * (1 - tp_factor)

            trade = {
                "cross_index": placed_index,
                "placed_time": str(placed_time),
                "placed_price": round2(placed_price),
                "side": side,
                "entry_time": time_str,
                "entry_price": round2(entry_price),
                "position_btc": position_btc,
                "lots": LOTS_PER_TRADE,
                "BTC_per_lot": BTC_PER_LOT,
                "leverage": LEVERAGE,
                "notional_usd": round(notional,2),
                "margin_used": round(margin_used,2),
                "SL": round2(sl),
                "TP": round2(tp),
                "TSL": round2(tsl),
                "highest": round2(highest) if side=="LONG" else None,
                "lowest": round2(lowest) if side=="SHORT" else None,
                "status": "OPEN",
                "exit_time": "",
                "exit_price": "",
                "exit_reason": "",
                "pnl_usd": "",
                "capital_before": round2(capital),
                "capital_after": round2(capital)
            }
            trades.append(trade)
            print(f"\n[ENTRY CONFIRMED] {side} executed at OPEN {entry_price:.2f} | position: {position_btc} BTC | margin: {margin_used:.2f} USD")
            print(f"   SL: {trade['SL']:.2f}  TP: {trade['TP']:.2f}  initial TSL: {trade['TSL']:.2f}\n")

            pending_order = None

        # 2) detect cross and place pending_order (execute next candle open)
        cross_event = None
        if prev is not None:
            if (prev['EMA_fast'] <= prev['EMA_slow']) and (last['EMA_fast'] > last['EMA_slow']):
                cross_event = {"index": len(candles)-1, "type":"GOLDEN", "time_bkk": last['candle_time_bkk'], "price": last['close']}
            elif (prev['EMA_fast'] >= prev['EMA_slow']) and (last['EMA_fast'] < last['EMA_slow']):
                cross_event = {"index": len(candles)-1, "type":"DEAD", "time_bkk": last['candle_time_bkk'], "price": last['close']}
            if cross_event:
                placed_side = "LONG" if cross_event['type']=="GOLDEN" else "SHORT"
                pending_order = {"side": placed_side, "placed_price": cross_event['price'], "placed_time": cross_event['time_bkk'], "placed_index": cross_event['index']}
                print(f"[SWING] {cross_event['type']} at {cross_event['time_bkk']} price {cross_event['price']:.2f}")
                print(f"[ORDER PLACED] Side: {placed_side} placed at cross-candle CLOSE {cross_event['price']:.2f} -> will execute next open\n")

        # 3) manage open trades: check TP, SL, TSL; update TSL only if price moves favorably
        for t in list(trades):
            if t['status'] != "OPEN":
                continue
            side = t['side']
            entry_p = float(t['entry_price'])
            pos_btc = float(t['position_btc'])

            if side == "LONG":
                # update highest & TSL
                if last['high'] > (t.get('highest') or entry_p):
                    t['highest'] = round2(last['high'])
                    t['TSL'] = round2(t['highest'] - TSL_DISTANCE)
                # TP
                if last['high'] >= float(t['TP']):
                    exit_price = float(t['TP'])
                    pnl = (exit_price - entry_p) * pos_btc
                    cap_before = capital
                    capital_change = pnl
                    capital += capital_change
                    t.update({"status":"CLOSED","exit_time": time_str,"exit_price": round2(exit_price),"exit_reason":"TP","pnl_usd": round2(pnl),"capital_before": round2(cap_before),"capital_after": round2(capital)})
                    print(f"[ORDER] LONG TP HIT at {exit_price:.2f} P&L: {pnl:.2f} NewCapital: {capital:.2f}")
                # SL or TSL
                elif last['low'] <= float(t['TSL']) or last['low'] <= float(t['SL']):
                    exit_price = float(t['TSL']) if last['low'] <= float(t['TSL']) else float(t['SL'])
                    pnl = (exit_price - entry_p) * pos_btc
                    cap_before = capital
                    capital += pnl
                    t.update({"status":"CLOSED","exit_time": time_str,"exit_price": round2(exit_price),"exit_reason":"TSL" if last['low'] <= float(t['TSL']) else "SL","pnl_usd": round2(pnl),"capital_before": round2(cap_before),"capital_after": round2(capital)})
                    print(f"[ORDER] LONG EXIT at {exit_price:.2f} due to {'TSL' if last['low'] <= float(t['TSL']) else 'SL'} P&L: {pnl:.2f} NewCapital: {capital:.2f}")

            else:  # SHORT
                if last['low'] < (t.get('lowest') or entry_p):
                    t['lowest'] = round2(last['low'])
                    t['TSL'] = round2(t['lowest'] + TSL_DISTANCE)
                if last['low'] <= float(t['TP']):
                    exit_price = float(t['TP'])
                    pnl = (entry_p - exit_price) * pos_btc
                    cap_before = capital
                    capital += pnl
                    t.update({"status":"CLOSED","exit_time": time_str,"exit_price": round2(exit_price),"exit_reason":"TP","pnl_usd": round2(pnl),"capital_before": round2(cap_before),"capital_after": round2(capital)})
                    print(f"[ORDER] SHORT TP HIT at {exit_price:.2f} P&L: {pnl:.2f} NewCapital: {capital:.2f}")
                elif last['high'] >= float(t['TSL']) or last['high'] >= float(t['SL']):
                    exit_price = float(t['TSL']) if last['high'] >= float(t['TSL']) else float(t['SL'])
                    pnl = (entry_p - exit_price) * pos_btc
                    cap_before = capital
                    capital += pnl
                    t.update({"status":"CLOSED","exit_time": time_str,"exit_price": round2(exit_price),"exit_reason":"TSL" if last['high'] >= float(t['TSL']) else "SL","pnl_usd": round2(pnl),"capital_before": round2(cap_before),"capital_after": round2(capital)})
                    print(f"[ORDER] SHORT EXIT at {exit_price:.2f} due to {'TSL' if last['high'] >= float(t['TSL']) else 'SL'} P&L: {pnl:.2f} NewCapital: {capital:.2f}")

        # 4) save to Excel (candles, historical swings/backtest, trades)
        try:
            write_df_safe(EXCEL_FILE, sheet_candles, candles)
            write_df_safe(EXCEL_FILE, sheet_swings, hist_trades_df)
            write_df_safe(EXCEL_FILE, sheet_trades, pd.DataFrame(trades))
        except Exception as ex:
            print("Excel write warning:", ex)

    except Exception as exc:
        print("Error in on_message:", exc)
        traceback.print_exc()

def on_open(ws):
    print("WebSocket opened.")

def on_error(ws, error):
    print("WebSocket error:", error)

def on_close(ws, close_status_code, close_msg):
    print("WebSocket closed.", close_status_code, close_msg)

# ========== start ==========
ws_url = f"wss://stream.binance.com:9443/ws/{SYMBOL.lower()}@kline_{INTERVAL}"
print(f"Starting live EMA {EMA_FAST}-{EMA_SLOW} strategy for {SYMBOL} with capital {capital_start:.2f} USD")
sheet_candles, sheet_swings, sheet_trades = ensure_excel_run_sheets(EXCEL_FILE, datetime.now(TZ_BKK).strftime("%Y%m%d_%H%M%S"))

ws = websocket.WebSocketApp(ws_url, on_open=on_open, on_message=on_message, on_error=on_error, on_close=on_close)
try:
    ws.run_forever()
except KeyboardInterrupt:
    print("Stopped by user.")
