# ema_trailing_live.py
import websocket
import json
import pandas as pd
import requests
from datetime import datetime
import pytz
import os
from openpyxl import load_workbook

# ---------------- USER PARAMETERS ----------------
symbol = "BTCUSDT"
interval = "1m"
history_limit = 500
stop_loss_pct = 0.02
target_pct = 0.035
trailing_stop_pct = 0.01  # 1% trailing stop
capital = 100000.0
risk_per_trade_pct = 0.01
ts = datetime.now().strftime("%Y%m%d_%H%M%S")
excel_file = f"excel_data/{symbol}_{ts}.xlsx"
tz_bkk = pytz.timezone("Asia/Bangkok")
# --------------------------------------------------

# in-memory storage
candles = pd.DataFrame(columns=['open','high','low','close','volume','candle_time_bkk','EMA9','EMA21'])
trades = []   # active and closed trades
swings = []   # EMA cross events

# ---------------- HELPER FUNCTIONS ----------------
def binance_klines(symbol, interval, limit=500):
    url = "https://api.binance.com/api/v3/klines"
    params = {"symbol": symbol, "interval": interval, "limit": limit}
    r = requests.get(url, params=params, timeout=10)
    r.raise_for_status()
    return r.json()

def to_bkk(ts_ms):
    return datetime.fromtimestamp(ts_ms/1000, pytz.utc).astimezone(tz_bkk)

def to_bkk_str(ts_ms):
    return to_bkk(ts_ms).strftime("%Y-%m-%d %H:%M:%S")

def calculate_emas(df):
    if len(df) >= 1:
        df['EMA9'] = df['close'].ewm(span=9, adjust=False).mean()
        df['EMA21'] = df['close'].ewm(span=21, adjust=False).mean()
    return df

def detect_crosses(df):
    crosses = []
    if 'EMA9' not in df or 'EMA21' not in df:
        return crosses
    for i in range(1, len(df)):
        a = df.iloc[i-1]
        b = df.iloc[i]
        if pd.isna(a['EMA9']) or pd.isna(a['EMA21']) or pd.isna(b['EMA9']) or pd.isna(b['EMA21']):
            continue
        if (a['EMA9'] <= a['EMA21']) and (b['EMA9'] > b['EMA21']):
            crosses.append({"index": i, "type": "GOLDEN", "time_bkk": b['candle_time_bkk'], "price": b['close']})
        if (a['EMA9'] >= a['EMA21']) and (b['EMA9'] < b['EMA21']):
            crosses.append({"index": i, "type": "DEAD", "time_bkk": b['candle_time_bkk'], "price": b['close']})
    return crosses

def backtest_historical(candles, swings, capital_start=100000.0):
    capital = capital_start
    risk_per_trade_pct = 0.01
    stop_loss_pct = 0.02
    target_pct = 0.035
    trailing_stop_pct = 0.01

    backtest_results = []

    for s in swings:
        entry_index = s['index']
        entry_price = candles.iloc[entry_index]['close']
        trade_type = "LONG" if s['type'] == "GOLDEN" else "SHORT"

        if trade_type == "LONG":
            sl = entry_price*(1-stop_loss_pct)
            tp = entry_price*(1+target_pct)
            tsl = entry_price*(1-trailing_stop_pct)
        else:
            sl = entry_price*(1+stop_loss_pct)
            tp = entry_price*(1-target_pct)
            tsl = entry_price*(1+trailing_stop_pct)

        # Track trade until exit
        for i in range(entry_index+1, len(candles)):
            candle = candles.iloc[i]
            exit_price = None
            exit_reason = None

            if trade_type == "LONG":
                tsl = max(tsl, candle['close']*(1-trailing_stop_pct))
                if candle['high'] >= tp:
                    exit_price = tp
                    exit_reason = "TP"
                elif candle['low'] <= sl or candle['low'] <= tsl:
                    exit_price = tsl if candle['low'] <= tsl else sl
                    exit_reason = "SL/TSL"
            else:
                tsl = min(tsl, candle['close']*(1+trailing_stop_pct))
                if candle['low'] <= tp:
                    exit_price = tp
                    exit_reason = "TP"
                elif candle['high'] >= sl or candle['high'] >= tsl:
                    exit_price = tsl if candle['high'] >= tsl else sl
                    exit_reason = "SL/TSL"

            if exit_price is not None:
                if trade_type == "LONG":
                    pnl_amount = capital * risk_per_trade_pct * ((exit_price-entry_price)/entry_price)
                else:
                    pnl_amount = capital * risk_per_trade_pct * ((entry_price-exit_price)/entry_price)
                capital_after = capital + pnl_amount

                backtest_results.append({
                    "entry_time": candle['candle_time_bkk'],
                    "entry_index": entry_index,
                    "entry_price": round(entry_price,2),
                    "type": trade_type,
                    "SL": round(sl,2),
                    "TP": round(tp,2),
                    "TSL": round(tsl,2),
                    "exit_time": candle['candle_time_bkk'],
                    "exit_price": round(exit_price,2),
                    "exit_reason": exit_reason,
                    "pnl_amount": round(pnl_amount,2),
                    "capital_before": round(capital,2),
                    "capital_after": round(capital_after,2)
                })
                capital = capital_after
                break

    return pd.DataFrame(backtest_results)

def ensure_excel_and_create_sheets(filename):
    if not os.path.exists(filename):
        writer = pd.ExcelWriter(filename, engine='openpyxl', mode='w')
        pd.DataFrame().to_excel(writer, sheet_name="Candles", index=False)
        pd.DataFrame().to_excel(writer, sheet_name="Swings", index=False)
        pd.DataFrame().to_excel(writer, sheet_name="Trades", index=False)
        pd.DataFrame().to_excel(writer, sheet_name="SwingsResult", index=False)
        writer.close()

def write_df_to_sheet(filename, sheet_name, df):
    df_copy = df.copy()
    # Remove timezone info for Excel
    for col in df_copy.columns:
        if pd.api.types.is_datetime64_any_dtype(df_copy[col]):
            df_copy[col] = df_copy[col].dt.tz_localize(None)
    with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_copy.to_excel(writer, sheet_name=sheet_name, index=False)

# ---------------- PRELOAD HISTORICAL CANDLES ----------------
print("Fetching historical candles...")
klines = binance_klines(symbol, interval, history_limit)
rows = []
for k in klines:
    rows.append({
        "open": round(float(k[1]), 2),
        "high": round(float(k[2]), 2),
        "low": round(float(k[3]), 2),
        "close": round(float(k[4]), 2),
        "volume": round(float(k[5]), 2),
        "candle_time_bkk": to_bkk(k[0])
    })
candles = pd.DataFrame(rows)
candles = calculate_emas(candles)
cross_events = detect_crosses(candles)

swings.extend(cross_events)
print(f"Loaded {len(candles)} historical candles, {len(cross_events)} cross events detected.")

# Excel setup
ensure_excel_and_create_sheets(excel_file)
write_df_to_sheet(excel_file, "Candles", candles)
write_df_to_sheet(excel_file, "Swings", pd.DataFrame(swings))
write_df_to_sheet(excel_file, "Trades", pd.DataFrame(trades))

# ---------------- Historical Backtest ----------------
historical_trades = backtest_historical(candles, cross_events, capital_start=100000.0)

# Save historical trades to Swings sheet
write_df_to_sheet(excel_file, "SwingsResult", historical_trades)
print("Historical trades and P&L saved to 'Swings' sheet.")

# ---------------- WEBSOCKET CALLBACKS ----------------
def on_message(ws, message):
    global candles, trades, capital, swings
    try:
        msg = json.loads(message)
        k = msg.get('k', {})
        if not k or not k.get('x', False):
            return

        t_open = int(k['t'])
        candle_time = to_bkk(t_open)
        new_row = {
            "open": round(float(k['o']),2),
            "high": round(float(k['h']),2),
            "low": round(float(k['l']),2),
            "close": round(float(k['c']),2),
            "volume": round(float(k['v']),2),
            "candle_time_bkk": candle_time
        }
        candles.loc[len(candles)] = new_row
        candles = calculate_emas(candles)

        last = candles.iloc[-1]
        prev = candles.iloc[-2] if len(candles)>=2 else None
        print(f"[{last['candle_time_bkk']}] Close: {last['close']} EMA9: {last['EMA9']:.2f} EMA21: {last['EMA21']:.2f}")

        # Detect cross
        cross_event = None
        if prev is not None:
            if prev['EMA9'] <= prev['EMA21'] and last['EMA9'] > last['EMA21']:
                cross_event = {"index": len(candles)-1, "type":"GOLDEN", "time_bkk": last['candle_time_bkk'], "price": last['close']}
            elif prev['EMA9'] >= prev['EMA21'] and last['EMA9'] < last['EMA21']:
                cross_event = {"index": len(candles)-1, "type":"DEAD", "time_bkk": last['candle_time_bkk'], "price": last['close']}
            if cross_event:
                swings.append(cross_event)
                print(f"[SWING] {cross_event['type']} CROSS at {cross_event['time_bkk']} price {cross_event['price']:.2f}")

        # Execute trade on cross
        if cross_event:
            entry_price = last['close']
            if cross_event['type'] == "GOLDEN":
                trade_type = "LONG"
                sl = entry_price*(1-stop_loss_pct)
                tp = entry_price*(1+target_pct)
                tsl = entry_price*(1-trailing_stop_pct)
            else:
                trade_type = "SHORT"
                sl = entry_price*(1+stop_loss_pct)
                tp = entry_price*(1-target_pct)
                tsl = entry_price*(1+trailing_stop_pct)

            trades.append({
                "type": trade_type,
                "entry_time": last['candle_time_bkk'],
                "entry_index": len(candles)-1,
                "entry_price": entry_price,
                "SL": round(sl,2),
                "TP": round(tp,2),
                "TSL": round(tsl,2),
                "status":"OPEN",
                "capital_before": capital,
                "capital_after": capital
            })
            print(f"[ORDER] {trade_type} ENTRY at {entry_price:.2f} SL:{sl:.2f} TP:{tp:.2f} TSL:{tsl:.2f}")

        # Update trades for TP/SL/TSL
        for t in trades:
            if t['status'] != 'OPEN':
                continue
            if t['type'] == 'LONG':
                # Update TSL if price moves up
                if last['close'] > t['entry_price']:
                    t['TSL'] = max(t['TSL'], round(last['close']*(1-trailing_stop_pct),2))
                # Check TP
                if last['high'] >= t['TP']:
                    pnl = capital * risk_per_trade_pct * target_pct
                    capital += pnl
                    t.update({"status":"CLOSED_TP","exit_time":last['candle_time_bkk'],"exit_price":t['TP'],"pnl_amount":round(pnl,2),"capital_after":capital})
                    print(f"[ORDER] LONG TP HIT at {t['TP']:.2f}, P&L:+{pnl:.2f}, New Capital:{capital:.2f}")
                # Check SL or TSL
                elif last['low'] <= t['SL'] or last['low'] <= t['TSL']:
                    exit_price = t['TSL'] if last['low'] <= t['TSL'] else t['SL']
                    pnl = capital * risk_per_trade_pct * ((exit_price - t['entry_price'])/t['entry_price'])
                    capital += pnl
                    t.update({"status":"CLOSED_SL","exit_time":last['candle_time_bkk'],"exit_price":exit_price,"pnl_amount":round(pnl,2),"capital_after":capital})
                    print(f"[ORDER] LONG EXIT at {exit_price:.2f} due to SL/TSL, P&L:{pnl:.2f}, New Capital:{capital:.2f}")

            else: # SHORT
                if last['close'] < t['entry_price']:
                    t['TSL'] = min(t['TSL'], round(last['close']*(1+trailing_stop_pct),2))
                if last['low'] <= t['TP']:
                    pnl = capital * risk_per_trade_pct * target_pct
                    capital += pnl
                    t.update({"status":"CLOSED_TP","exit_time":last['candle_time_bkk'],"exit_price":t['TP'],"pnl_amount":round(pnl,2),"capital_after":capital})
                    print(f"[ORDER] SHORT TP HIT at {t['TP']:.2f}, P&L:+{pnl:.2f}, New Capital:{capital:.2f}")
                elif last['high'] >= t['SL'] or last['high'] >= t['TSL']:
                    exit_price = t['TSL'] if last['high'] >= t['TSL'] else t['SL']
                    pnl = capital * risk_per_trade_pct * ((t['entry_price']-exit_price)/t['entry_price'])
                    capital += pnl
                    t.update({"status":"CLOSED_SL","exit_time":last['candle_time_bkk'],"exit_price":exit_price,"pnl_amount":round(pnl,2),"capital_after":capital})
                    print(f"[ORDER] SHORT EXIT at {exit_price:.2f} due to SL/TSL, P&L:{pnl:.2f}, New Capital:{capital:.2f}")

        # Save to Excel
        write_df_to_sheet(excel_file, "Candles", candles)
        write_df_to_sheet(excel_file, "Swings", pd.DataFrame(swings))
        write_df_to_sheet(excel_file, "Trades", pd.DataFrame(trades))

    except Exception as e:
        print("Error in on_message:", e)

def on_open(ws):
    print("WebSocket opened.")

def on_error(ws, error):
    print("WebSocket error:", error)

def on_close(ws, close_status_code, close_msg):
    print("WebSocket closed.", close_status_code, close_msg)

# ---------------- RUN WEBSOCKET ----------------
ws_url = f"wss://stream.binance.com:9443/ws/{symbol.lower()}@kline_{interval}"
print(f"Starting live EMA 9-21 strategy for {symbol} with capital {capital}")
ws = websocket.WebSocketApp(ws_url, on_open=on_open, on_message=on_message, on_error=on_error, on_close=on_close)
ws.run_forever()
