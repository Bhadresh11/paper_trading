import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from datetime import datetime

class ExcelManager:
    def __init__(self, symbol):
        folder = "excel_data"
        os.makedirs(folder, exist_ok=True)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.file_path = f"{folder}/{symbol}_{ts}.xlsx"

        # Create workbook with sheets
        if not os.path.exists(self.file_path):
            wb = Workbook()

            # Default sheet → rename to Candles
            ws = wb.active
            ws.title = "Candles"

            wb.create_sheet("Swings")
            wb.create_sheet("Orders")
            wb.create_sheet("TSL_Updates")
            wb.create_sheet("Stats")

            wb.save(self.file_path)

    # ----------- INTERNAL SAVE (replace sheet) -----------
    def _write_sheet(self, sheet_name, df):
        """Write full dataframe to sheet"""
        with pd.ExcelWriter(self.file_path, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # ----------- INTERNAL APPEND -----------
    def _append_sheet(self, sheet_name, df_new):
        """Append rows to sheet while preserving old data."""

        try:
            old = pd.read_excel(self.file_path, sheet_name=sheet_name)
            df_final = pd.concat([old, df_new], ignore_index=True)
        except Exception:
            df_final = df_new

        self._write_sheet(sheet_name, df_final)

    # ----------- SAVE METHODS -----------

    def save_candle(self, c):
        df = pd.DataFrame([c])
        self._append_sheet("Candles", df)

    def save_candles_bulk(self, candles):
        """Save historical candles faster"""
        df = pd.DataFrame(candles)
        self._append_sheet("Candles", df)

    # In data_manager.py

    def save_swing(self, high_utc, high_local, high_value, low_utc, low_local, low_value):
        """
        Save swing levels in Excel with both UTC and Local times
        """
        new_row = {
            "swing_high_utc": high_utc,
            "swing_high_local": high_local,
            "swing_high_value": high_value,
            "swing_low_utc": low_utc,
            "swing_low_local": low_local,
            "swing_low_value": low_value
        }
        # Initialize if empty
        if not hasattr(self, 'swings_df') or self.swings_df is None:
            self.swings_df = pd.DataFrame(columns=new_row.keys())

        self.swings_df = pd.concat([self.swings_df, pd.DataFrame([new_row])], ignore_index=True)
        self._save_to_excel("Swings", self.swings_df)


    def save_order(self, o):
        df = pd.DataFrame([o])
        self._append_sheet("Orders", df)

    def save_tsl_update(self, t):
        df = pd.DataFrame([t])
        self._append_sheet("TSL_Updates", df)

    def save_stats(self, s):
        df = pd.DataFrame([s])
        self._append_sheet("Stats", df)
