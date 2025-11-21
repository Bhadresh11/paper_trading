import asyncio
import json
import requests
import websockets
import datetime
import openpyxl
from openpyxl import Workbook
import pytz
import os

# -------------------------------------
# CONFIG
# -------------------------------------
SYMBOL = "ethusdt"
INTERVAL = "1m"
EXCEL_FILE = "ethusdt_data.xlsx"
LOOKBACK = 15

local_tz = pytz.timezone("Asia/Phnom_Penh")

# GLOBALS
historical_loaded = False
candles = []
last_swing_high = None
last_swing_low = None


# ================================================================
# EXCEL FUNCTIONS
# ================================================================
def init_excel():
    """Create Excel file or load existing."""
    if os.path.exists(EXCEL_FILE):
        print("[EXCEL] Loaded existing file.")
        return openpyxl.load_workbook(EXCEL_FILE)

    print("[EXCEL] New Excel file created.")
    wb = Workbook()

    # Candle sheet
    s1 = wb.active
    s1.title = "ETHUSD"
    s1.append(["Candle Time", "Local Time", "Open", "High", "Low", "Close", "Volume"])

    # Swing sheet
    s2 = wb.create_sheet("SWINGS")
    s2.append(["Candle Time", "Local Time", "Swing High", "Swing Low"])

    wb.save(EXCEL_FILE)
    return wb


def save_candle(wb, candle):
    sheet = wb["ETHUSD"]

    # Avoid duplicate save
    if sheet.max_row > 1:
        last_row = sheet[sheet.max_row]
        if str(last_row[0].value) == candle["candle_time"]:
            return

    sheet.append([
        candle["candle_time"],
        candle["local_time"],
        candle["open"],
        candle["high"],
        candle["low"],
        candle["close"],
        candle["volume"]
    ])
    wb.save(EXCEL_FILE)

    print(f"[EXCEL] Saved candle {candle['candle_time']}")


def save_swing(wb, candle_time, local_time, high, low):
    sheet = wb["SWINGS"]

    sheet.append([candle_time, local_time, high, low])
    wb.save(EXCEL_FILE)

    print(f"[EXCEL] Saved SWING: High={high}, Low={low}")


# ================================================================
# SWING LOGIC
# ================================================================
def calculate_swing(candle_list):
    if len(candle_list) < LOOKBACK:
        return None, None

    highs = [c["high"] for c in candle_list[-LOOKBACK:]]
    lows = [c["low"] for c in candle_list[-LOOKBACK:]]

    return max(highs), min(lows)


# ================================================================
# FETCH HISTORICAL DATA (ONLY ONCE)
# ================================================================
def load_historical():
    global historical_loaded, candles

    if historical_loaded:
        return

    url = (
        f"https://api.binance.com/api/v3/klines?"
        f"symbol={SYMBOL.upper()}&interval={INTERVAL}&limit=500"
    )
    print("[API] Loading historical candles…")

    res = requests.get(url)

    if res.status_code != 200:
        print(f"[API ERROR] {res.status_code}: {res.text}")
        return

    raw = res.json()

    for c in raw:
        candle_time = datetime.datetime.fromtimestamp(c[0] / 1000, pytz.UTC)
        local_tm = candle_time.astimezone(local_tz).strftime("%H:%M:%S")

        candles.append({
            "candle_time": candle_time.strftime("%Y-%m-%d %H:%M:%S"),
            "local_time": local_tm,
            "open": float(c[1]),
            "high": float(c[2]),
            "low": float(c[3]),
            "close": float(c[4]),
            "volume": float(c[5]),
        })

    historical_loaded = True
    print(f"[API] Historical load done: {len(candles)} candles")


# ================================================================
# WEBSOCKET LIVE DATA
# ================================================================
async def websocket_loop():
    ws_url = f"wss://stream.binance.com:9443/ws/{SYMBOL}@kline_{INTERVAL}"

    wb = init_excel()
    load_historical()

    # Save historical candles only once
    print("[HIST] Writing historical candles to Excel…")
    for c in candles:
        save_candle(wb, c)

    # Compute initial swing
    global last_swing_high, last_swing_low
    last_swing_high, last_swing_low = calculate_swing(candles)
    print(f"[INIT SWING] High={last_swing_high}, Low={last_swing_low}")

    save_swing(
        wb,
        candles[-1]["candle_time"],
        candles[-1]["local_time"],
        last_swing_high,
        last_swing_low
    )

    # --------------- WebSocket loop ---------------
    while True:
        try:
            print("[WS] Connecting to Binance…")

            async with websockets.connect(
                ws_url,
                ping_interval=20,
                ping_timeout=20,
                close_timeout=5
            ) as ws:

                print("[WS] Connected!")

                while True:
                    msg = await ws.recv()
                    data = json.loads(msg)
                    k = data["k"]

                    candle_time = datetime.datetime.fromtimestamp(k["t"] / 1000, pytz.UTC)
                    local_tm = candle_time.astimezone(local_tz).strftime("%H:%M:%S")

                    # live tick log
                    print(f"[TICK] {local_tm} Close={k['c']}")

                    # Only process complete candles
                    if k["x"] is False:
                        continue

                    # Build new candle
                    new_candle = {
                        "candle_time": candle_time.strftime("%Y-%m-%d %H:%M:%S"),
                        "local_time": local_tm,
                        "open": float(k["o"]),
                        "high": float(k["h"]),
                        "low": float(k["l"]),
                        "close": float(k["c"]),
                        "volume": float(k["v"]),
                    }

                    # Append & save
                    candles.append(new_candle)
                    save_candle(wb, new_candle)

                    # Recalculate swing
                    swing_high, swing_low = calculate_swing(candles)

                    # Save swing ONLY IF CHANGED
                    global last_swing_high, last_swing_low
                    if swing_high != last_swing_high or swing_low != last_swing_low:
                        print(f"[NEW SWING] High={swing_high}, Low={swing_low}")
                        save_swing(
                            wb,
                            new_candle["candle_time"],
                            new_candle["local_time"],
                            swing_high,
                            swing_low
                        )
                        last_swing_high = swing_high
                        last_swing_low = swing_low

        except Exception as e:
            print(f"[WS ERROR] {e}")
            print("[WS] Reconnecting in 3 sec…")
            await asyncio.sleep(3)


asyncio.run(websocket_loop())
