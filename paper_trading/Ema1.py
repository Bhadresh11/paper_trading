import websocket
import json
import pandas as pd
import requests
from datetime import datetime
import pytz
import os
from openpyxl import load_workbook

# ---------------- USER SETTINGS ----------------
symbol = "BTCUSDT"
interval = "1m"
history_limit = 500
stop_loss_pct = 0.02
target_pct = 0.035
capital = 100000.0
risk_per_trade_pct = 0.01
ts = datetime.now().strftime("%Y%m%d_%H%M%S")
excel_file = f"excel_data/{symbol}_{ts}.xlsx"

tz_bkk = pytz.timezone("Asia/Bangkok")

# ---------------- MEMORY ----------------
candles = pd.DataFrame(columns=[
    'open','high','low','close','volume','time_bkk','EMA9','EMA21'
])
trades = []     # executed trades
swings = []     # detected cross events
pending_order = None   # holds order placed at cross; to be executed at next candle open

# ---------------- HELPERS ----------------
def to_bkk_str(ts_ms):
    t = datetime.fromtimestamp(ts_ms / 1000, pytz.utc).astimezone(tz_bkk)
    return t.strftime("%Y-%m-%d %H:%M:%S")

def binance_klines(symbol, interval, limit):
    url = "https://api.binance.com/api/v3/klines"
    r = requests.get(url, params={"symbol": symbol, "interval": interval, "limit": limit}, timeout=10)
    r.raise_for_status()
    return r.json()

def calculate_emas(df):
    if len(df) > 1:
        df["EMA9"] = df["close"].ewm(span=9, adjust=False).mean()
        df["EMA21"] = df["close"].ewm(span=21, adjust=False).mean()
    return df

def detect_crosses(df):
    results = []
    for i in range(1, len(df)):
        a = df.iloc[i - 1]
        b = df.iloc[i]
        if pd.isna(a.get("EMA9")) or pd.isna(a.get("EMA21")):
            continue
        if a["EMA9"] <= a["EMA21"] and b["EMA9"] > b["EMA21"]:
            results.append({"index": i, "type": "GOLDEN", "time_bkk": b["time_bkk"], "price": b["close"]})
        if a["EMA9"] >= a["EMA21"] and b["EMA9"] < b["EMA21"]:
            results.append({"index": i, "type": "DEAD", "time_bkk": b["time_bkk"], "price": b["close"]})
    return results

def ensure_excel_sheets():
    sheets = {"candles": "Candles", "Historical_Trades": "Historical_Trades", "swings": "Swings", "orders": "Orders"}
    if not os.path.exists(excel_file):
        writer = pd.ExcelWriter(excel_file, engine="openpyxl", mode="w")
        pd.DataFrame().to_excel(writer, sheet_name=sheets["candles"], index=False)
        pd.DataFrame().to_excel(writer, sheet_name=sheets["Historical_Trades"], index=False)
        pd.DataFrame().to_excel(writer, sheet_name=sheets["swings"], index=False)
        pd.DataFrame().to_excel(writer, sheet_name=sheets["orders"], index=False)
        writer.close()
    else:
        # add new sheets (unique)
        wb = load_workbook(excel_file)
        def unique_name(name):
            if name not in wb.sheetnames:
                return name
            n = 1
            while f"{name}_{n}" in wb.sheetnames:
                n += 1
            return f"{name}_{n}"
        sheets = {k: unique_name(v) for k,v in sheets.items()}
        writer = pd.ExcelWriter(excel_file, engine="openpyxl", mode="a")
        pd.DataFrame().to_excel(writer, sheet_name=sheets["candles"], index=False)
        pd.DataFrame().to_excel(writer, sheet_name=sheets["Historical_Trades"], index=False)
        pd.DataFrame().to_excel(writer, sheet_name=sheets["swings"], index=False)
        pd.DataFrame().to_excel(writer, sheet_name=sheets["orders"], index=False)
        writer.close()
    return sheets

def calculate_historical_trades(candles, crosses, stop_loss_pct, target_pct, starting_capital=100000):
    trades = []
    capital = starting_capital
    n = len(candles)

    for i, cross in enumerate(crosses):

        trade_type = "LONG" if cross["type"] == "GOLDEN" else "SHORT"
        entry_index = cross["index"]
        entry_price = candles.iloc[entry_index]["close"]
        entry_time = candles.iloc[entry_index]["time_bkk"]

        # SL and TP
        if trade_type == "LONG":
            SL = entry_price * (1 - stop_loss_pct)
            TP = entry_price * (1 + target_pct)
        else:
            SL = entry_price * (1 + stop_loss_pct)
            TP = entry_price * (1 - target_pct)

        exit_price = None
        exit_time = None
        exit_reason = None

        # search for exit
        for j in range(entry_index + 1, n):
            row = candles.iloc[j]
            high = row["high"]
            low = row["low"]

            # LONG EXIT LOGIC
            if trade_type == "LONG":
                if high >= TP:
                    exit_price = TP
                    exit_time = row["time_bkk"]
                    exit_reason = "TP"
                    break
                if low <= SL:
                    exit_price = SL
                    exit_time = row["time_bkk"]
                    exit_reason = "SL"
                    break

            # SHORT EXIT LOGIC
            else:
                if low <= TP:
                    exit_price = TP
                    exit_time = row["time_bkk"]
                    exit_reason = "TP"
                    break
                if high >= SL:
                    exit_price = SL
                    exit_time = row["time_bkk"]
                    exit_reason = "SL"
                    break

            # Exit on opposite cross
            if i + 1 < len(crosses) and crosses[i + 1]["index"] == j:
                exit_price = row["close"]
                exit_time = row["time_bkk"]
                exit_reason = "OppositeCross"
                break

        # If no exit found → exit at the last candle
        if exit_price is None:
            last_row = candles.iloc[-1]
            exit_price = last_row["close"]
            exit_time = last_row["time_bkk"]
            exit_reason = "EndOfHistory"

        # Calculate P&L
        if trade_type == "LONG":
            pnl_pct = (exit_price - entry_price) / entry_price
        else:
            pnl_pct = (entry_price - exit_price) / entry_price

        # Capital updates
        capital_before = capital
        pnl_amount = capital * pnl_pct
        capital_after = capital_before + pnl_amount

        # Store trade record
        trades.append({
            "entry_time": entry_time,
            "entry_index": entry_index,
            "entry_price": round(entry_price, 2),
            "type": trade_type,
            "SL": round(SL, 2),
            "TP": round(TP, 2),
            "exit_time": exit_time,
            "exit_price": round(exit_price, 2),
            "exit_reason": exit_reason,

            # PNL info
            "pnl_pct": round(pnl_pct * 100, 2),       # as percentage
            "pnl_amount": round(pnl_amount, 2),
            "capital_before": round(capital_before, 2),
            "capital_after": round(capital_after, 2)
        })

        # Update capital for next trade
        capital = capital_after

    return pd.DataFrame(trades)

def write_df(sheet, df):
    with pd.ExcelWriter(excel_file, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet, index=False)

# ---------- Load historical candles ----------
print("Loading historical candles...")
klines = binance_klines(symbol, interval, history_limit)
rows = []
for k in klines:
    rows.append({
        "open": round(float(k[1]), 2),
        "high": round(float(k[2]), 2),
        "low": round(float(k[3]), 2),
        "close": round(float(k[4]), 2),
        "volume":round(float(k[5]), 2),
        "time_bkk": to_bkk_str(k[0])
    })
candles = pd.DataFrame(rows)
candles = calculate_emas(candles)
crosses = detect_crosses(candles)
if crosses:
    print("Last historical cross:", crosses[-1])
else:
    print("No historical cross found.")

sheets = ensure_excel_sheets()
write_df(sheets["candles"], candles)
write_df(sheets["swings"], pd.DataFrame(crosses))
print("Initial data written to Excel sheets:", sheets)
history_trades = calculate_historical_trades(
    candles=candles,
    crosses=crosses,
    stop_loss_pct=stop_loss_pct,
    target_pct=target_pct,
    starting_capital=100000
)

print(history_trades)
write_df(sheets["Historical_Trades"], history_trades)

# ---------------- WEBSOCKET CALLBACKS ----------------
def on_message(ws, message):
    global candles, trades, swings, pending_order, capital

    try:
        msg = json.loads(message)
        k = msg.get("k", {})
        if not k or not k.get("x", False):
            return

        t_open_ms = int(k["t"])
        time_bkk = to_bkk_str(t_open_ms)
        # Build the closed candle row (we get full OHLC for the candle that just closed)
        new_row = {
            "open": float(k["o"]),
            "high": float(k["h"]),
            "low": float(k["l"]),
            "close": float(k["c"]),
            "volume": float(k["v"]),
            "time_bkk": time_bkk
        }

        # Append new candle (loc to avoid concat warning)
        candles.loc[len(candles)] = new_row
        candles = calculate_emas(candles)
        last = candles.iloc[-1]
        prev = candles.iloc[-2] if len(candles) > 1 else None

        # Print candle close + EMAs
        print(f"[{last['time_bkk']}] Close: {last['close']:.2f}  EMA9: {last['EMA9']:.2f} EMA21: {last['EMA21']:.2f}")

        # ------------- Pending order execution at this candle's OPEN -------------
        # If we have a pending_order placed on previous candle close, we execute it now at THIS candle's OPEN price.
        if pending_order is not None:
            # Use this candle's open as execution price
            exec_price = new_row["open"]
            method = pending_order["placed_at"]['method']  # 'CLOSE' (placed at close)
            placed_time = pending_order["placed_at"]['time_bkk']
            placed_price = pending_order["placed_at"]['price']
            side = pending_order["side"]  # 'LONG' or 'SHORT'

            # build executed trade record
            if side == "LONG":
                sl = exec_price * (1 - stop_loss_pct)
                tp = exec_price * (1 + target_pct)
            else:
                sl = exec_price * (1 + stop_loss_pct)
                tp = exec_price * (1 - target_pct)

            trade_rec = {
                "placed_time_bkk": placed_time,
                "placed_price": placed_price,
                "placed_method": method,
                "entry_time_bkk": time_bkk,
                "entry_price": exec_price,
                "entry_method": "OPEN_EXECUTED",
                "type": side,
                "SL": sl,
                "TP": tp,
                "status": "OPEN",
                "exit_time_bkk": "",
                "pnl": "",
                "capital_before": capital,
                "pnl_live": 0.0,      # new live P&L field
            }
            trades.append(trade_rec)

            # print ENTRY CONFIRMED block (exact details)
            print("\n[ENTRY CONFIRMED]")
            print(f"Placed at (cross): {placed_time} price={placed_price:.2f} method={method}")
            print(f"Executed at next candle OPEN: {time_bkk} price={exec_price:.2f}")
            print(f"Side: {side}, SL: {sl:.2f}, TP: {tp:.2f}, Capital before: {capital:.2f}\n")

            # clear pending order
            pending_order = None

        # ---------------- Detect new EMA cross and PLACE ORDER at this candle (will execute next candle open) ---------------
        if prev is not None:
            # GOLDEN CROSS detected on this closed candle -> PLACE an order at cross-candle close price (will execute next open)
            if prev["EMA9"] <= prev["EMA21"] and last["EMA9"] > last["EMA21"]:
                placed = {"time_bkk": last["time_bkk"], "price": last["close"], "method": "CLOSE"}
                side = "LONG"
                pending_order = {"placed_at": placed, "side": side}
                swings.append({"type": "GOLDEN", "time_bkk": last["time_bkk"], "price": last["close"]})
                print(f"[SWING] GOLDEN CROSS at {last['time_bkk']} price {last['close']:.2f}")
                print(f"[ORDER PLACED] Side: {side}, placed_at_close_price: {last['close']:.2f} (will execute at next candle open)")

            # DEAD CROSS -> PLACE SHORT order
            elif prev["EMA9"] >= prev["EMA21"] and last["EMA9"] < last["EMA21"]:
                placed = {"time_bkk": last["time_bkk"], "price": last["close"], "method": "CLOSE"}
                side = "SHORT"
                pending_order = {"placed_at": placed, "side": side}
                swings.append({"type": "DEAD", "time_bkk": last["time_bkk"], "price": last["close"]})
                print(f"[SWING] DEAD CROSS at {last['time_bkk']} price {last['close']:.2f}")
                print(f"[ORDER PLACED] Side: {side}, placed_at_close_price: {last['close']:.2f} (will execute at next candle open)")

        # ---------------- Check open trades for TP/SL on this candle's high/low ----------------
        for t in trades:
            if t["status"] != "OPEN":
                entry = t["entry_price"]
                cur = last["close"]

                # Long position P&L
                if t["type"] == "LONG":
                    pnl_live = cur - entry

                # Short position P&L
                else:
                    pnl_live = entry - cur

                # store live P&L (but not the final SL/TP P&L)
                t["pnl_live"] = pnl_live

                print(f"[LIVE PNL] {t['type']} | Entry: {entry:.2f} | Current: {cur:.2f} | PNL: {pnl_live:.2f}")
                continue
            if t["type"] == "LONG":
                # TP hit
                if last["high"] >= t["TP"]:
                    pnl = capital * risk_per_trade_pct * target_pct
                    capital += pnl
                    t["status"] = "CLOSED_TP"
                    t["exit_time_bkk"] = last["time_bkk"]
                    t["pnl"] = pnl
                    print(f"[ORDER] LONG TP HIT at {t['TP']:.2f} P&L:+{pnl:.2f} NewCap:{capital:.2f}")
                # SL hit
                elif last["low"] <= t["SL"]:
                    pnl = -capital * risk_per_trade_pct * stop_loss_pct
                    capital += pnl
                    t["status"] = "CLOSED_SL"
                    t["exit_time_bkk"] = last["time_bkk"]
                    t["pnl"] = pnl
                    print(f"[ORDER] LONG SL HIT at {t['SL']:.2f} P&L:{pnl:.2f} NewCap:{capital:.2f}")

            else:  # SHORT
                if last["low"] <= t["TP"]:
                    pnl = capital * risk_per_trade_pct * target_pct
                    capital += pnl
                    t["status"] = "CLOSED_TP"
                    t["exit_time_bkk"] = last["time_bkk"]
                    t["pnl"] = pnl
                    print(f"[ORDER] SHORT TP HIT at {t['TP']:.2f} P&L:+{pnl:.2f} NewCap:{capital:.2f}")
                elif last["high"] >= t["SL"]:
                    pnl = -capital * risk_per_trade_pct * stop_loss_pct
                    capital += pnl
                    t["status"] = "CLOSED_SL"
                    t["exit_time_bkk"] = last["time_bkk"]
                    t["pnl"] = pnl
                    print(f"[ORDER] SHORT SL HIT at {t['SL']:.2f} P&L:{pnl:.2f} NewCap:{capital:.2f}")

        # ---------------- Write updates to Excel (safe) ----------------
        try:
            write_df(sheets["candles"], candles)
        except Exception as e:
            print("Warning: failed writing candles sheet:", e)

        try:
            swings_df = pd.DataFrame(swings)
            write_df(sheets["swings"], swings_df if not swings_df.empty else pd.DataFrame())
        except Exception as e:
            print("Warning: failed writing swings sheet:", e)

        try:
            trades_df = pd.DataFrame(trades)
            if "pnl_live" not in trades_df.columns:
                trades_df["pnl_live"] = ""
                    
            write_df(sheets["orders"], trades_df)
            # write_df(sheets["orders"], trades_df if not trades_df.empty else pd.DataFrame())
        except Exception as e:
            print("Warning: failed writing orders sheet:", e)

    except Exception as exc:
        print("Error in on_message:", exc)

def on_open(ws):
    print("WebSocket Connected")

def on_error(ws, error):
    print("WebSocket ERROR:", error)

def on_close(ws, a, b):
    print("WebSocket Closed")

# ---------------- START ----------------
print(f"Starting EMA 9–21 Algo for {symbol} | Capital: {capital}")
# Prepare Excel sheets for this run
# sheets = ensure_excel_sheets()
# write initial history
write_df(sheets["candles"], candles)
write_df(sheets["swings"], pd.DataFrame(crosses) if (crosses := detect_crosses(candles)) else pd.DataFrame())

ws_url = f"wss://stream.binance.com:9443/ws/{symbol.lower()}@kline_{interval}"
ws = websocket.WebSocketApp(ws_url, on_open=on_open, on_message=on_message, on_error=on_error, on_close=on_close)
ws.run_forever()
